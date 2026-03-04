"""
PROMEOS - Routes Patrimoine (WORLD CLASS)
VNext pipeline: template, import, quality gate, corrections, activation, export.
CRUD Sites/Compteurs/Contrats + QA scoring.
"""

import csv
import io
import json
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Any, Dict, Optional, List
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from database import get_db
from models import (
    Organisation,
    EntiteJuridique,
    Portefeuille,
    StagingBatch,
    StagingSite,
    StagingCompteur,
    QualityFinding,
    ImportSourceType,
    StagingStatus,
    QualityRuleSeverity,
    ActivationLog,
    ActivationLogStatus,
    Site,
    DeliveryPoint,
    not_deleted,
    Compteur,
    TypeSite,
    TypeCompteur,
    EnergyVector,
    EnergyContract,
    BillingEnergyType,
    StatutConformite,
    PaymentRule,
    PaymentRuleLevel,
    ContractIndexation,
    ContractStatus,
)
from services.patrimoine_service import (
    create_staging_batch,
    import_csv_to_staging,
    import_invoices_to_staging,
    get_staging_summary,
    run_quality_gate,
    apply_fix,
    activate_batch,
    get_diff_plan,
    compute_content_hash,
    abandon_batch,
)
from services.import_mapping import (
    CANONICAL_COLUMNS,
    generate_csv_template,
    generate_xlsx_template,
    map_headers,
    detect_encoding,
    detect_delimiter,
    normalize_column_name,
    get_mapping_report,
)
from middleware.auth import get_optional_auth, get_portfolio_optional_auth, AuthContext
from services.scope_utils import get_scope_org_id
from routes.billing import check_contract_overlap

router = APIRouter(prefix="/api/patrimoine", tags=["Patrimoine"])


# ========================================
# Schemas
# ========================================


class FixRequest(BaseModel):
    fix_type: str
    params: dict


class BulkFixRequest(BaseModel):
    fixes: List[FixRequest]


class ActivateRequest(BaseModel):
    portefeuille_id: int


class InvoiceImportRequest(BaseModel):
    invoices: list


class UpdateFieldRequest(BaseModel):
    staging_site_id: Optional[int] = None
    staging_compteur_id: Optional[int] = None
    field: str
    value: Optional[str] = None


class SiteUpdateRequest(BaseModel):
    nom: Optional[str] = None
    adresse: Optional[str] = None
    code_postal: Optional[str] = None
    ville: Optional[str] = None
    region: Optional[str] = None
    surface_m2: Optional[float] = None
    nombre_employes: Optional[int] = None
    naf_code: Optional[str] = None
    siret: Optional[str] = None
    type: Optional[str] = None


class SiteMergeRequest(BaseModel):
    source_site_id: int
    target_site_id: int


class CompteurMoveRequest(BaseModel):
    target_site_id: int


class CompteurUpdateRequest(BaseModel):
    numero_serie: Optional[str] = None
    meter_id: Optional[str] = None
    puissance_souscrite_kw: Optional[float] = None
    type: Optional[str] = None


class ContractCreateRequest(BaseModel):
    site_id: int
    energy_type: str = "elec"
    supplier_name: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    price_ref_eur_per_kwh: Optional[float] = None
    fixed_fee_eur_per_month: Optional[float] = None
    notice_period_days: int = 90
    auto_renew: bool = False
    # V96
    offer_indexation: Optional[str] = None
    price_granularity: Optional[str] = None
    renewal_alert_days: Optional[int] = None
    contract_status: Optional[str] = None


class ContractUpdateRequest(BaseModel):
    supplier_name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    price_ref_eur_per_kwh: Optional[float] = None
    fixed_fee_eur_per_month: Optional[float] = None
    notice_period_days: Optional[int] = None
    auto_renew: Optional[bool] = None
    # V96
    offer_indexation: Optional[str] = None
    price_granularity: Optional[str] = None
    renewal_alert_days: Optional[int] = None
    contract_status: Optional[str] = None


# V96: Payment Rules schemas
class PaymentRuleCreateRequest(BaseModel):
    level: str  # portefeuille | site | contrat
    portefeuille_id: Optional[int] = None
    site_id: Optional[int] = None
    contract_id: Optional[int] = None
    invoice_entity_id: int
    payer_entity_id: Optional[int] = None
    cost_center: Optional[str] = None


class PaymentRuleBulkApplyRequest(BaseModel):
    site_ids: List[int]
    invoice_entity_id: int
    payer_entity_id: Optional[int] = None
    cost_center: Optional[str] = None


# V97: Resolution Engine schemas
class ReconciliationFixRequest(BaseModel):
    action: str
    params: Optional[Dict[str, Any]] = None


# ========================================
# Multi-org scope helpers
# ========================================


def _get_org_id(request: Request, auth: Optional[AuthContext], db: Session) -> int:
    """Resolve org_id via centralized scope chain (DEMO_MODE-aware)."""
    from services.scope_utils import resolve_org_id

    return resolve_org_id(request, auth, db)


def _check_batch_org(batch: StagingBatch, org_id: int):
    """Verify batch belongs to the resolved org. Raises 403 if mismatch."""
    if batch is None:
        raise HTTPException(status_code=404, detail="Batch non trouvé")
    if batch.org_id is not None and batch.org_id != org_id:
        raise HTTPException(status_code=403, detail="Batch hors périmètre")


def _check_site_belongs_to_org(db: Session, site: Site, org_id: int):
    """Verify site belongs to org via portfolio→EJ chain. Fail-closed: raises 403 on any break."""
    if not site.portefeuille_id:
        raise HTTPException(status_code=403, detail="Site hors périmètre")
    pf = db.query(Portefeuille).get(site.portefeuille_id)
    if not pf:
        raise HTTPException(status_code=403, detail="Site hors périmètre")
    ej = db.query(EntiteJuridique).get(pf.entite_juridique_id)
    if not ej or ej.organisation_id != org_id:
        raise HTTPException(status_code=403, detail="Site hors périmètre")


def _check_portfolio_belongs_to_org(db: Session, portfolio_id: int, org_id: int):
    """Verify portfolio belongs to org. Raises 403 if mismatch."""
    pf = db.query(Portefeuille).get(portfolio_id)
    if not pf:
        raise HTTPException(status_code=404, detail=f"Portefeuille {portfolio_id} non trouvé")
    ej = db.query(EntiteJuridique).get(pf.entite_juridique_id)
    if not ej or ej.organisation_id != org_id:
        raise HTTPException(status_code=403, detail="Portefeuille hors périmètre")
    return pf


def _load_site_with_org_check(db: Session, site_id: int, org_id: int) -> Site:
    """Load a site and verify org ownership. Raises 404/403."""
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail=f"Site {site_id} non trouvé")
    _check_site_belongs_to_org(db, site, org_id)
    return site


def _load_compteur_with_org_check(db: Session, compteur_id: int, org_id: int) -> "Compteur":
    """Load a compteur with upfront org verification via JOIN chain. Returns 404 on miss."""
    c = (
        db.query(Compteur)
        .join(Site, Compteur.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(Compteur.id == compteur_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail=f"Compteur {compteur_id} non trouvé")
    return c


def _load_contract_with_org_check(db: Session, contract_id: int, org_id: int) -> "EnergyContract":
    """Load a contract with upfront org verification via JOIN chain. Returns 404 on miss."""
    ct = (
        db.query(EnergyContract)
        .join(Site, EnergyContract.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(EnergyContract.id == contract_id)
        .first()
    )
    if not ct:
        raise HTTPException(status_code=404, detail=f"Contrat {contract_id} non trouvé")
    return ct


# ========================================
# Template download
# ========================================


@router.get("/import/template")
def import_template(
    format: str = Query("xlsx", description="xlsx or csv"),
):
    """Download official patrimoine import template."""
    if format == "csv":
        content = generate_csv_template()
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment; filename=template_patrimoine.csv"},
        )
    else:
        try:
            content = generate_xlsx_template()
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed — Excel template unavailable")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_patrimoine.xlsx"},
        )


@router.get("/import/template/columns")
def import_template_columns():
    """List canonical template columns with metadata."""
    return {
        "columns": CANONICAL_COLUMNS,
        "delimiter": ";",
        "encoding": "utf-8",
        "notes": [
            "Seule la colonne 'nom' est obligatoire.",
            "Les noms de colonnes sont flexibles (synonymes auto-detectes).",
            "delivery_code = PRM (elec) ou PCE (gaz), 14 chiffres.",
        ],
    }


# ========================================
# Import endpoints
# ========================================


@router.post("/staging/import")
async def staging_import(
    request: Request,
    file: UploadFile = File(...),
    mode: str = Query("import", description="express, import, assiste, demo"),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Import CSV/Excel file into staging pipeline.

    Performs: encoding detection, header mapping, normalization, dedup check.
    """
    org_id = _get_org_id(request, auth, db)

    content = await file.read()
    content_hash = compute_content_hash(content)

    # Detect source type
    filename = file.filename or ""
    if filename.endswith((".xlsx", ".xls")):
        source_type = ImportSourceType.EXCEL
    else:
        source_type = ImportSourceType.CSV

    # Check for duplicate import (same content hash, same org)
    existing = (
        db.query(StagingBatch)
        .filter(
            StagingBatch.content_hash == content_hash,
            StagingBatch.org_id == org_id,
            StagingBatch.status != StagingStatus.ABANDONED,
        )
        .first()
    )
    if existing:
        summary = get_staging_summary(db, existing.id)
        return {
            "batch_id": existing.id,
            "duplicate": True,
            "detail": "File already imported",
            **summary,
        }

    batch = create_staging_batch(
        db=db,
        org_id=org_id,
        user_id=auth.user.id if auth else None,
        source_type=source_type,
        mode=mode,
        filename=filename,
        content_hash=content_hash,
    )

    # Detect header mapping for the response
    mapping_info = None
    if source_type == ImportSourceType.EXCEL:
        try:
            result = _parse_excel_to_staging(db, batch.id, content)
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl not installed — Excel import unavailable")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel parse error: {e}")
    else:
        # Detect mapping from CSV headers
        encoding = detect_encoding(content)
        text = content.decode(encoding)
        first_line = text.split("\n")[0].strip()
        delimiter = detect_delimiter(first_line)
        raw_headers = [h.strip() for h in first_line.split(delimiter)]
        header_mapping, mapping_warnings = map_headers(raw_headers)
        mapping_info = {
            "mapping": {k: v for k, v in header_mapping.items() if k != v},
            "warnings": mapping_warnings,
            "encoding": encoding,
            "delimiter": delimiter,
        }
        result = import_csv_to_staging(db, batch.id, content)

    db.commit()

    response = {
        "batch_id": batch.id,
        "duplicate": False,
        **result,
    }
    if mapping_info:
        response["mapping"] = mapping_info

    return response


@router.post("/staging/import-invoices")
def staging_import_invoices(
    request: Request,
    body: InvoiceImportRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Import sites/meters from invoice metadata into staging."""
    org_id = _get_org_id(request, auth, db)

    batch = create_staging_batch(
        db=db,
        org_id=org_id,
        user_id=auth.user.id if auth else None,
        source_type=ImportSourceType.INVOICE,
        mode="assiste",
    )

    result = import_invoices_to_staging(db, batch.id, {"invoices": body.invoices})
    db.commit()

    return {
        "batch_id": batch.id,
        **result,
    }


# ========================================
# Quality gate & corrections
# ========================================


@router.get("/staging/{batch_id}/summary")
def staging_summary(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Get staging batch summary stats."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    try:
        return get_staging_summary(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/staging/{batch_id}/rows")
def staging_rows(
    batch_id: int,
    request: Request,
    status: Optional[str] = Query(None, description="ok, error, skipped"),
    q: Optional[str] = Query(None, description="Search query"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List staging rows (sites + linked compteurs) with pagination & search."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)

    query = db.query(StagingSite).filter(StagingSite.batch_id == batch_id)

    # Filter by status
    if status == "skipped":
        query = query.filter(StagingSite.skip.is_(True))
    elif status == "ok":
        query = query.filter(StagingSite.skip.is_(False))
    elif status == "error":
        # Sites with unresolved findings
        error_site_ids = (
            db.query(QualityFinding.staging_site_id)
            .filter(
                QualityFinding.batch_id == batch_id,
                QualityFinding.resolved.is_(False),
                QualityFinding.staging_site_id.isnot(None),
            )
            .distinct()
            .all()
        )
        error_ids = [r[0] for r in error_site_ids]
        query = query.filter(StagingSite.id.in_(error_ids)) if error_ids else query.filter(False)

    # Search
    if q:
        search = f"%{q}%"
        query = query.filter(
            (StagingSite.nom.ilike(search))
            | (StagingSite.adresse.ilike(search))
            | (StagingSite.ville.ilike(search))
            | (StagingSite.siret.ilike(search))
        )

    total = query.count()
    sites = query.order_by(StagingSite.row_number).offset((page - 1) * page_size).limit(page_size).all()

    # Get findings indexed by staging_site_id
    all_findings = (
        db.query(QualityFinding)
        .filter(
            QualityFinding.batch_id == batch_id,
            QualityFinding.resolved.is_(False),
        )
        .all()
    )
    site_findings = {}
    for f in all_findings:
        if f.staging_site_id:
            site_findings.setdefault(f.staging_site_id, []).append(f)

    rows = []
    for ss in sites:
        compteurs = (
            db.query(StagingCompteur)
            .filter(
                StagingCompteur.staging_site_id == ss.id,
            )
            .all()
        )

        findings_for_site = site_findings.get(ss.id, [])

        rows.append(
            {
                "id": ss.id,
                "row_number": ss.row_number,
                "nom": ss.nom,
                "adresse": ss.adresse,
                "code_postal": ss.code_postal,
                "ville": ss.ville,
                "surface_m2": ss.surface_m2,
                "type_site": ss.type_site,
                "siret": ss.siret,
                "naf_code": ss.naf_code,
                "skip": ss.skip,
                "target_site_id": ss.target_site_id,
                "issues_count": len(findings_for_site),
                "compteurs": [
                    {
                        "id": sc.id,
                        "row_number": sc.row_number,
                        "numero_serie": sc.numero_serie,
                        "meter_id": sc.meter_id,
                        "type_compteur": sc.type_compteur,
                        "puissance_kw": sc.puissance_kw,
                        "skip": sc.skip,
                    }
                    for sc in compteurs
                ],
            }
        )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "rows": rows,
    }


@router.get("/staging/{batch_id}/issues")
def staging_issues(
    batch_id: int,
    request: Request,
    severity: Optional[str] = Query(None, description="blocking, critical, warning, info"),
    resolved: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List quality findings (issues) for a batch, optionally filtered."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)

    query = db.query(QualityFinding).filter(QualityFinding.batch_id == batch_id)

    if severity:
        try:
            sev_enum = QualityRuleSeverity(severity)
            query = query.filter(QualityFinding.severity == sev_enum)
        except ValueError:
            pass

    if resolved is not None:
        query = query.filter(QualityFinding.resolved.is_(resolved))

    findings = query.order_by(QualityFinding.id).all()

    return {
        "total": len(findings),
        "issues": [
            {
                "id": f.id,
                "rule_id": f.rule_id,
                "severity": f.severity.value,
                "staging_site_id": f.staging_site_id,
                "staging_compteur_id": f.staging_compteur_id,
                "evidence": f.evidence_json,
                "suggested_action": f.suggested_action,
                "resolved": f.resolved,
                "resolution": f.resolution,
            }
            for f in findings
        ],
    }


@router.post("/staging/{batch_id}/validate")
def staging_validate(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Run quality gate on staging batch."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    try:
        findings = run_quality_gate(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    blocking_count = sum(1 for f in findings if f["severity"] in ("blocking", "critical"))
    db.commit()

    return {
        "findings": findings,
        "blocking_count": blocking_count,
        "can_activate": blocking_count == 0,
    }


@router.put("/staging/{batch_id}/fix")
def staging_fix(
    batch_id: int,
    request: Request,
    body: FixRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Apply a correction to staging data."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    result = apply_fix(db, batch_id, body.fix_type, body.params)
    db.commit()
    return result


@router.put("/staging/{batch_id}/fix/bulk")
def staging_fix_bulk(
    batch_id: int,
    request: Request,
    body: BulkFixRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Apply multiple corrections in a single transaction."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    results = []
    for fix in body.fixes:
        r = apply_fix(db, batch_id, fix.fix_type, fix.params)
        results.append(r)
    db.commit()
    applied_count = sum(1 for r in results if r.get("applied"))
    return {
        "applied": applied_count,
        "total": len(results),
        "results": results,
    }


@router.post("/staging/{batch_id}/autofix")
def staging_autofix(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Apply safe auto-corrections to staging data.

    Safe fixes:
    - Trim whitespace on all text fields
    - Pad code_postal to 5 digits
    - Normalize type_compteur (electricite/gaz/eau)
    - Skip orphan compteurs without meter_id and without numero_serie
    """
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)

    fixes_applied = 0

    # Fix 1: Trim + normalize staging sites
    sites = (
        db.query(StagingSite)
        .filter(
            StagingSite.batch_id == batch_id,
            StagingSite.skip.is_(False),
        )
        .all()
    )

    for ss in sites:
        changed = False
        # Trim whitespace
        for field in ("nom", "adresse", "ville", "siret", "naf_code"):
            val = getattr(ss, field, None)
            if val and val != val.strip():
                setattr(ss, field, val.strip())
                changed = True
        # Pad postal code
        if ss.code_postal:
            padded = ss.code_postal.strip().zfill(5)[:5]
            if padded != ss.code_postal:
                ss.code_postal = padded
                changed = True
        if changed:
            fixes_applied += 1

    # Fix 2: Normalize compteur types + trim
    compteurs = (
        db.query(StagingCompteur)
        .filter(
            StagingCompteur.batch_id == batch_id,
            StagingCompteur.skip.is_(False),
        )
        .all()
    )

    for sc in compteurs:
        changed = False
        # Trim meter_id
        if sc.meter_id and sc.meter_id != sc.meter_id.strip():
            sc.meter_id = sc.meter_id.strip()
            changed = True
        # Normalize type_compteur
        if sc.type_compteur:
            normalized = _normalize_compteur_type(sc.type_compteur)
            if normalized != sc.type_compteur:
                sc.type_compteur = normalized
                changed = True
        # Skip empty compteurs (no meter_id AND no numero_serie)
        if not sc.meter_id and not sc.numero_serie:
            sc.skip = True
            changed = True
        if changed:
            fixes_applied += 1

    db.flush()
    db.commit()

    return {
        "fixes_applied": fixes_applied,
        "detail": f"Applied {fixes_applied} safe auto-corrections",
    }


@router.delete("/staging/{batch_id}")
def staging_abandon(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Abandon a staging batch."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    try:
        result = abandon_batch(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    db.commit()
    return result


# ========================================
# Activation
# ========================================


@router.post("/staging/{batch_id}/activate")
def staging_activate(
    batch_id: int,
    request: Request,
    body: ActivateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Activate a validated staging batch → create real entities."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)
    _check_portfolio_belongs_to_org(db, body.portefeuille_id, org_id)
    try:
        result = activate_batch(db, batch_id, body.portefeuille_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    db.commit()
    return result


@router.get("/staging/{batch_id}/result")
def staging_result(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Get activation result for a batch (post-activation)."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)

    summary = get_staging_summary(db, batch_id)

    # Find activation log
    log = (
        db.query(ActivationLog)
        .filter(
            ActivationLog.batch_id == batch_id,
        )
        .order_by(ActivationLog.id.desc())
        .first()
    )

    result = {
        "batch_id": batch_id,
        "status": batch.status.value if batch.status else None,
        "mode": batch.mode,
        "filename": batch.filename,
        **summary,
    }

    if log:
        result["activation"] = {
            "log_id": log.id,
            "status": log.status.value,
            "started_at": log.started_at.isoformat() if log.started_at else None,
            "completed_at": log.completed_at.isoformat() if log.completed_at else None,
            "sites_created": log.sites_created,
            "compteurs_created": log.compteurs_created,
            "error_message": log.error_message,
        }

    # Stats from batch
    if batch.stats_json:
        try:
            stats = json.loads(batch.stats_json)
            result["stats"] = stats
        except (json.JSONDecodeError, TypeError):
            pass

    return result


@router.get("/staging/{batch_id}/export/report.csv")
def staging_export_report(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Export batch report as CSV: all rows + issues + status."""
    org_id = _get_org_id(request, auth, db)
    batch = db.query(StagingBatch).get(batch_id)
    _check_batch_org(batch, org_id)

    sites = (
        db.query(StagingSite)
        .filter(
            StagingSite.batch_id == batch_id,
        )
        .order_by(StagingSite.row_number)
        .all()
    )

    findings = (
        db.query(QualityFinding)
        .filter(
            QualityFinding.batch_id == batch_id,
        )
        .all()
    )

    # Index findings by staging_site_id
    site_findings_map = {}
    for f in findings:
        if f.staging_site_id:
            site_findings_map.setdefault(f.staging_site_id, []).append(f)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    # Header
    writer.writerow(
        [
            "row",
            "nom",
            "adresse",
            "code_postal",
            "ville",
            "surface_m2",
            "type_site",
            "siret",
            "naf_code",
            "status",
            "compteurs",
            "issues_count",
            "issues_detail",
        ]
    )

    for ss in sites:
        compteurs = (
            db.query(StagingCompteur)
            .filter(
                StagingCompteur.staging_site_id == ss.id,
            )
            .all()
        )
        cpt_list = "; ".join(f"{sc.meter_id or sc.numero_serie or '?'} ({sc.type_compteur or '?'})" for sc in compteurs)

        findings_for_site = site_findings_map.get(ss.id, [])
        issues_detail = "; ".join(f"[{f.severity.value}] {f.rule_id}" for f in findings_for_site)

        status = "skipped" if ss.skip else ("merged" if ss.target_site_id else "active")

        writer.writerow(
            [
                ss.row_number,
                ss.nom,
                ss.adresse,
                ss.code_postal,
                ss.ville,
                ss.surface_m2,
                ss.type_site,
                ss.siret,
                ss.naf_code,
                status,
                cpt_list,
                len(findings_for_site),
                issues_detail,
            ]
        )

    csv_bytes = output.getvalue().encode("utf-8-sig")

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename=rapport_import_batch_{batch_id}.csv"},
    )


# ========================================
# Delivery Points
# ========================================


@router.get("/sites/{site_id}/delivery-points")
def site_delivery_points(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List active delivery points (PRM/PCE) for a site."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, site_id, org_id)

    dps = (
        not_deleted(db.query(DeliveryPoint), DeliveryPoint)
        .filter(
            DeliveryPoint.site_id == site_id,
        )
        .all()
    )

    return [
        {
            "id": dp.id,
            "code": dp.code,
            "energy_type": dp.energy_type.value if dp.energy_type else None,
            "status": dp.status.value if dp.status else None,
            "compteurs_count": len(dp.compteurs) if dp.compteurs else 0,
            "data_source": dp.data_source,
            "created_at": dp.created_at.isoformat() if dp.created_at else None,
        }
        for dp in dps
    ]


# ========================================
# Incremental sync
# ========================================


@router.post("/{portfolio_id}/sync")
async def portfolio_sync(
    portfolio_id: int,
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Incremental sync: compare uploaded file vs existing portfolio."""
    org_id = _get_org_id(request, auth, db)
    _check_portfolio_belongs_to_org(db, portfolio_id, org_id)

    content = await file.read()
    content_hash = compute_content_hash(content)
    filename = file.filename or ""

    # Create temporary staging batch
    batch = create_staging_batch(
        db=db,
        org_id=org_id,
        user_id=auth.user.id if auth else None,
        source_type=ImportSourceType.CSV,
        mode="sync",
        filename=filename,
        content_hash=content_hash,
    )
    import_csv_to_staging(db, batch.id, content)

    # Get diff plan
    diff = get_diff_plan(db, portfolio_id, batch.id)

    if not dry_run:
        # Apply: activate new sites
        result = activate_batch(db, batch.id, portfolio_id)
        diff["applied"] = True
        diff["activation"] = result
    else:
        diff["applied"] = False

    db.commit()
    return diff


# ========================================
# Demo loader
# ========================================


@router.post("/demo/load")
def demo_load(db: Session = Depends(get_db)):
    """Load demo patrimoine data (Collectivite Azur). Requires DEMO_MODE."""
    from middleware.auth import DEMO_MODE

    if not DEMO_MODE:
        raise HTTPException(status_code=403, detail="Demo load désactivé en production")
    try:
        from scripts.seed_data import seed_patrimoine_demo

        result = seed_patrimoine_demo(db)
        db.commit()
        return {"status": "ok", **result}
    except ImportError:
        raise HTTPException(status_code=500, detail="seed_data module not available")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# Helpers
# ========================================


def _normalize_compteur_type(raw: str) -> str:
    """Normalize compteur type string for autofix."""
    low = raw.lower().strip()
    if any(k in low for k in ("elec", "elect")):
        return "electricite"
    if any(k in low for k in ("gaz", "gas")):
        return "gaz"
    if any(k in low for k in ("eau", "water")):
        return "eau"
    return raw


def _parse_excel_to_staging(db: Session, batch_id: int, content: bytes) -> dict:
    """Parse Excel file via openpyxl and feed into staging."""
    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"sites_count": 0, "compteurs_count": 0, "parse_errors": [{"row": 0, "error": "Empty workbook"}]}

    # First row = headers — normalize via mapping
    raw_headers = [str(h or "").strip() for h in rows[0]]
    normalized_headers = [normalize_column_name(h) for h in raw_headers]

    # Convert to CSV bytes with normalized headers
    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(normalized_headers)
    for row in rows[1:]:
        writer.writerow([str(c) if c is not None else "" for c in row])

    csv_bytes = output.getvalue().encode("utf-8")
    return import_csv_to_staging(db, batch_id, csv_bytes)


# ========================================
# Mapping preview (header recognition)
# ========================================


class MappingPreviewRequest(BaseModel):
    headers: list


@router.post("/mapping/preview")
def mapping_preview(body: MappingPreviewRequest):
    """Preview how CSV/Excel headers will be mapped to canonical columns."""
    return get_mapping_report(body.headers)


# ========================================
# KPIs (server-side aggregation)
# ========================================


@router.get("/kpis")
def patrimoine_kpis(
    request: Request,
    site_id: Optional[int] = Query(None, description="Filter KPIs to a single site"),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Aggregated KPIs for the patrimoine page — replaces client-side useMemo."""
    org_id = _get_org_id(request, auth, db)

    base_q = (
        db.query(Site)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(Site.actif == True)
    )

    if site_id is not None:
        base_q = base_q.filter(Site.id == site_id)

    result = base_q.with_entities(
        func.count(Site.id).label("total"),
        func.count(case((Site.statut_decret_tertiaire == StatutConformite.CONFORME, 1))).label("conformes"),
        func.count(case((Site.statut_decret_tertiaire == StatutConformite.A_RISQUE, 1))).label("a_risque"),
        func.count(case((Site.statut_decret_tertiaire == StatutConformite.NON_CONFORME, 1))).label("non_conformes"),
        func.coalesce(func.sum(Site.risque_financier_euro), 0).label("total_risque"),
        func.coalesce(func.sum(Site.surface_m2), 0).label("total_surface"),
        func.count(case((Site.anomalie_facture == True, 1))).label("total_anomalies"),
    ).one()

    return {
        "total": result.total,
        "conformes": result.conformes,
        "aRisque": result.a_risque,
        "nonConformes": result.non_conformes,
        "totalRisque": round(float(result.total_risque), 2),
        "totalSurface": round(float(result.total_surface), 2),
        "totalAnomalies": result.total_anomalies,
    }


# ========================================
# Site CRUD (WORLD CLASS)
# ========================================


def _serialize_site(site: Site) -> dict:
    return {
        "id": site.id,
        "nom": site.nom,
        "type": site.type.value if site.type else None,
        "adresse": site.adresse,
        "code_postal": site.code_postal,
        "ville": site.ville,
        "region": site.region,
        "surface_m2": site.surface_m2,
        "nombre_employes": site.nombre_employes,
        "siret": site.siret,
        "naf_code": site.naf_code,
        "actif": site.actif,
        "portefeuille_id": site.portefeuille_id,
        "data_source": site.data_source,
        "created_at": site.created_at.isoformat() if site.created_at else None,
        "updated_at": site.updated_at.isoformat() if site.updated_at else None,
        # Enriched analytics fields
        "risque_eur": site.risque_financier_euro,
        "statut_conformite": site.statut_decret_tertiaire.value if site.statut_decret_tertiaire else None,
        "anomalie_facture": site.anomalie_facture,
        "conso_kwh_an": site.annual_kwh_total,
    }


def _build_sites_query(
    db: Session, org_id: int, portefeuille_id=None, actif=None, ville=None, type_site=None, search=None
):
    """Build a filtered site query scoped to org — shared by list_sites and export."""
    q = (
        db.query(Site)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
    )
    if portefeuille_id is not None:
        q = q.filter(Site.portefeuille_id == portefeuille_id)
    if actif is not None:
        q = q.filter(Site.actif == actif)
    if ville:
        q = q.filter(Site.ville.ilike(f"%{ville}%"))
    if type_site:
        q = q.filter(Site.type == type_site)
    if search:
        q = q.filter(
            (Site.nom.ilike(f"%{search}%")) | (Site.ville.ilike(f"%{search}%")) | (Site.adresse.ilike(f"%{search}%"))
        )
    return q


_SORT_WHITELIST = {"nom", "ville", "surface_m2", "risque_financier_euro", "type", "created_at"}


@router.get("/sites")
def list_sites(
    request: Request,
    portefeuille_id: Optional[int] = None,
    actif: Optional[bool] = None,
    ville: Optional[str] = None,
    type_site: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(25, ge=1, le=200, description="Items per page"),
    sort_by: Optional[str] = Query(None, description="Sort column"),
    sort_dir: Optional[str] = Query("asc", description="Sort direction: asc or desc"),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List sites with filters, pagination, and sorting — scoped to org."""
    org_id = _get_org_id(request, auth, db)
    q = _build_sites_query(db, org_id, portefeuille_id, actif, ville, type_site, search)

    # Sort
    if sort_by and sort_by in _SORT_WHITELIST:
        col = getattr(Site, sort_by, None)
        if col is not None:
            q = q.order_by(col.desc() if sort_dir == "desc" else col.asc())

    total = q.count()

    # Use page/page_size if page > 1, otherwise fall back to skip/limit for backward compat
    if page > 1 or page_size != 25:
        offset = (page - 1) * page_size
        sites = q.offset(offset).limit(page_size).all()
    else:
        sites = q.offset(skip).limit(limit).all()

    return {
        "total": total,
        "sites": [_serialize_site(s) for s in sites],
        "page": page,
        "page_size": page_size,
    }


@router.get("/sites/export.csv")
def export_sites_csv(
    request: Request,
    portefeuille_id: Optional[int] = None,
    actif: Optional[bool] = None,
    ville: Optional[str] = None,
    type_site: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Export filtered sites as CSV (streaming, UTF-8-sig BOM for French Excel)."""
    org_id = _get_org_id(request, auth, db)
    q = _build_sites_query(db, org_id, portefeuille_id, actif, ville, type_site, search)
    sites = q.all()

    headers = [
        "id",
        "nom",
        "type",
        "adresse",
        "code_postal",
        "ville",
        "region",
        "surface_m2",
        "nombre_employes",
        "siret",
        "actif",
        "risque_financier_euro",
        "statut_conformite",
        "anomalie_facture",
        "conso_kwh_an",
        "portefeuille_id",
    ]

    def iter_csv():
        yield "\ufeff"  # BOM for Excel
        out = io.StringIO()
        w = csv.writer(out, delimiter=";")
        w.writerow(headers)
        yield out.getvalue()
        for site in sites:
            out = io.StringIO()
            w = csv.writer(out, delimiter=";")
            w.writerow(
                [
                    site.id,
                    site.nom,
                    site.type.value if site.type else "",
                    site.adresse or "",
                    site.code_postal or "",
                    site.ville or "",
                    site.region or "",
                    site.surface_m2 or "",
                    site.nombre_employes or "",
                    site.siret or "",
                    site.actif,
                    site.risque_financier_euro or 0,
                    site.statut_decret_tertiaire.value if site.statut_decret_tertiaire else "",
                    site.anomalie_facture or False,
                    site.annual_kwh_total or "",
                    site.portefeuille_id or "",
                ]
            )
            yield out.getvalue()

    filename = f"patrimoine_sites_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter_csv(),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sites/{site_id}")
def get_site_detail(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Get a site with compteurs and contracts count."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, site_id, org_id)
    compteurs_count = db.query(Compteur).filter(Compteur.site_id == site_id, Compteur.actif.is_(True)).count()
    contracts_count = db.query(EnergyContract).filter(EnergyContract.site_id == site_id).count()
    return {
        **_serialize_site(site),
        "compteurs_count": compteurs_count,
        "contracts_count": contracts_count,
    }


@router.patch("/sites/{site_id}")
def update_site(
    site_id: int,
    request: Request,
    body: SiteUpdateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Update a site (partial update)."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, site_id, org_id)

    updated_fields = []
    for field, value in body.model_dump(exclude_unset=True).items():
        if field == "type" and value is not None:
            try:
                value = TypeSite(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Type invalide: {value}")
        setattr(site, field, value)
        updated_fields.append(field)

    db.commit()
    return {"updated": updated_fields, **_serialize_site(site)}


@router.post("/sites/{site_id}/archive")
def archive_site(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Soft-delete a site (set actif=False)."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, site_id, org_id)
    if not site.actif:
        return {"detail": "Site deja archive", "site_id": site_id}
    site.actif = False
    db.commit()
    return {"detail": "Site archive", "site_id": site_id}


@router.post("/sites/{site_id}/restore")
def restore_site(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Restore an archived site (set actif=True)."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, site_id, org_id)
    if site.actif:
        return {"detail": "Site deja actif", "site_id": site_id}
    site.actif = True
    db.commit()
    return {"detail": "Site restaure", "site_id": site_id}


@router.post("/sites/merge")
def merge_sites(
    request: Request,
    body: SiteMergeRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Merge source site into target: transfer compteurs+contracts, archive source."""
    org_id = _get_org_id(request, auth, db)
    source = _load_site_with_org_check(db, body.source_site_id, org_id)
    target = _load_site_with_org_check(db, body.target_site_id, org_id)
    if source.id == target.id:
        raise HTTPException(status_code=400, detail="Source et cible identiques")

    # Transfer compteurs
    compteurs_moved = (
        db.query(Compteur)
        .filter(Compteur.site_id == source.id)
        .update({"site_id": target.id}, synchronize_session="fetch")
    )
    # Transfer contracts
    contracts_moved = (
        db.query(EnergyContract)
        .filter(EnergyContract.site_id == source.id)
        .update({"site_id": target.id}, synchronize_session="fetch")
    )
    # Archive source
    source.actif = False
    db.commit()

    return {
        "detail": f"Site {source.id} fusionne dans {target.id}",
        "compteurs_moved": compteurs_moved,
        "contracts_moved": contracts_moved,
        "source_archived": True,
    }


# ========================================
# Compteur Operations (WORLD CLASS)
# ========================================


def _serialize_compteur(c: Compteur) -> dict:
    return {
        "id": c.id,
        "site_id": c.site_id,
        "type": c.type.value if c.type else None,
        "numero_serie": c.numero_serie,
        "meter_id": c.meter_id,
        "puissance_souscrite_kw": c.puissance_souscrite_kw,
        "energy_vector": c.energy_vector.value if c.energy_vector else None,
        "actif": c.actif,
        "data_source": c.data_source,
    }


@router.get("/compteurs")
def list_compteurs(
    request: Request,
    site_id: Optional[int] = None,
    actif: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List compteurs with filters — scoped to org."""
    org_id = _get_org_id(request, auth, db)
    q = (
        db.query(Compteur)
        .join(Site, Compteur.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
    )
    if site_id is not None:
        q = q.filter(Compteur.site_id == site_id)
    if actif is not None:
        q = q.filter(Compteur.actif == actif)
    total = q.count()
    compteurs = q.offset(skip).limit(limit).all()
    return {"total": total, "compteurs": [_serialize_compteur(c) for c in compteurs]}


@router.patch("/compteurs/{compteur_id}")
def update_compteur(
    compteur_id: int,
    request: Request,
    body: CompteurUpdateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Update a compteur (partial update)."""
    org_id = _get_org_id(request, auth, db)
    c = _load_compteur_with_org_check(db, compteur_id, org_id)

    updated = []
    for field, value in body.model_dump(exclude_unset=True).items():
        if field == "type" and value is not None:
            try:
                value = TypeCompteur(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Type invalide: {value}")
        setattr(c, field, value)
        updated.append(field)

    db.commit()
    return {"updated": updated, **_serialize_compteur(c)}


@router.post("/compteurs/{compteur_id}/move")
def move_compteur(
    compteur_id: int,
    request: Request,
    body: CompteurMoveRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Move a compteur to another site."""
    org_id = _get_org_id(request, auth, db)
    c = _load_compteur_with_org_check(db, compteur_id, org_id)
    target = _load_site_with_org_check(db, body.target_site_id, org_id)

    old_site_id = c.site_id
    c.site_id = target.id
    db.commit()
    return {
        "detail": f"Compteur {compteur_id} deplace de site {old_site_id} vers {target.id}",
        **_serialize_compteur(c),
    }


@router.post("/compteurs/{compteur_id}/detach")
def detach_compteur(
    compteur_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Deactivate a compteur (soft detach)."""
    org_id = _get_org_id(request, auth, db)
    c = _load_compteur_with_org_check(db, compteur_id, org_id)
    c.actif = False
    db.commit()
    return {"detail": f"Compteur {compteur_id} desactive", **_serialize_compteur(c)}


# ========================================
# Contract CRUD (WORLD CLASS)
# ========================================


def _serialize_contract(ct: EnergyContract) -> dict:
    return {
        "id": ct.id,
        "site_id": ct.site_id,
        "energy_type": ct.energy_type.value if ct.energy_type else None,
        "supplier_name": ct.supplier_name,
        "start_date": ct.start_date.isoformat() if ct.start_date else None,
        "end_date": ct.end_date.isoformat() if ct.end_date else None,
        "price_ref_eur_per_kwh": ct.price_ref_eur_per_kwh,
        "fixed_fee_eur_per_month": ct.fixed_fee_eur_per_month,
        "notice_period_days": ct.notice_period_days,
        "auto_renew": ct.auto_renew,
        # V96
        "offer_indexation": ct.offer_indexation.value if ct.offer_indexation else None,
        "price_granularity": ct.price_granularity,
        "renewal_alert_days": ct.renewal_alert_days,
        "contract_status": ct.contract_status.value if ct.contract_status else None,
        "created_at": ct.created_at.isoformat() if ct.created_at else None,
    }


# ========================================
# Response models — Snapshot & Anomalies (V59)
# ========================================


class RegulatoryImpact(BaseModel):
    framework: str  # DECRET_TERTIAIRE / FACTURATION / BACS / NONE
    risk_level: str  # HIGH / MEDIUM / LOW
    explanation_fr: str


class BusinessImpact(BaseModel):
    type: str  # DATA_QUALITY / REGULATORY_RISK / BILLING_RISK
    estimated_risk_eur: float
    confidence: float  # 0..1
    explanation_fr: str


class AnomalyResponse(BaseModel):
    code: str
    severity: str
    title_fr: str
    detail_fr: str
    evidence: Dict[str, Any]
    cta: Dict[str, str]
    fix_hint_fr: str
    # V59 additions (always present, null-safe)
    regulatory_impact: Optional[RegulatoryImpact] = None
    business_impact: Optional[BusinessImpact] = None
    priority_score: Optional[int] = None


class SiteAnomaliesResponse(BaseModel):
    site_id: int
    anomalies: List[AnomalyResponse]
    completude_score: int
    nb_anomalies: int
    computed_at: str
    # V59 additions
    total_estimated_risk_eur: float
    assumptions_used: Dict[str, Any]


class OrgAnomaliesSiteItem(BaseModel):
    site_id: int
    nom: str
    completude_score: int
    nb_anomalies: int
    top_severity: Optional[str]
    top_priority_score: Optional[int]
    total_estimated_risk_eur: float
    anomalies: List[AnomalyResponse]


class OrgAnomaliesResponse(BaseModel):
    total: int
    page: int
    page_size: int
    sites: List[OrgAnomaliesSiteItem]


# ── V60/V61 : Portfolio summary ────────────────────────────────────────────


class PortfolioSitesAtRisk(BaseModel):
    critical: int = 0
    high: int = 0
    medium: int = 0
    low: int = 0


class PortfolioSitesHealth(BaseModel):
    """V61 — distribution des sites par score de complétude (data quality)."""

    healthy: int = 0  # completude_score >= 85
    warning: int = 0  # 50 <= completude_score < 85
    critical: int = 0  # completude_score < 50
    healthy_pct: float = 0.0


class PortfolioTrend(BaseModel):
    """V61 — tendance vs snapshot précédent. Null si pas d'historique."""

    risk_eur_delta: Optional[float] = None
    sites_count_delta: Optional[int] = None
    direction: Optional[str] = None  # "up" | "down" | "stable" | null
    vs_computed_at: Optional[str] = None


class PortfolioFrameworkItem(BaseModel):
    framework: str
    risk_eur: float
    anomalies_count: int


class PortfolioTopSiteItem(BaseModel):
    site_id: int
    site_nom: str
    risk_eur: float
    anomalies_count: int
    top_framework: Optional[str] = None


class PortfolioSummaryResponse(BaseModel):
    scope: Dict[str, Any]
    total_estimated_risk_eur: float
    sites_count: int
    sites_at_risk: PortfolioSitesAtRisk
    sites_health: PortfolioSitesHealth  # V61 NEW
    framework_breakdown: List[PortfolioFrameworkItem]
    top_sites: List[PortfolioTopSiteItem]
    trend: Optional[PortfolioTrend] = None  # V61 NEW (null — pas d'historique encore)
    computed_at: str


# ========================================
# Snapshot & Anomalies (V58 → V59)
# ========================================


@router.get("/sites/{site_id}/snapshot")
def get_site_snapshot_endpoint(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """
    Snapshot canonique d'un site : surface SoT, bâtiments, compteurs,
    points de livraison, contrats.  Scoped org — zéro N+1.
    """
    from services.patrimoine_snapshot import get_site_snapshot

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)  # 404/403 si hors périmètre
    snapshot = get_site_snapshot(site_id, org_id, db)
    if snapshot is None:
        raise HTTPException(status_code=404, detail=f"Site {site_id} non trouvé")
    return snapshot


@router.get("/sites/{site_id}/anomalies", response_model=SiteAnomaliesResponse)
def get_site_anomalies_endpoint(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """
    Anomalies de données patrimoine pour un site (8 règles P0).
    V59 : enrichies avec regulatory_impact, business_impact, priority_score.
    Triées par priority_score DESC.
    """
    from services.patrimoine_anomalies import compute_site_anomalies
    from services.patrimoine_impact import enrich_anomalies_with_impact
    from services.patrimoine_snapshot import get_site_snapshot
    from config.patrimoine_assumptions import DEFAULT_ASSUMPTIONS

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)

    result = compute_site_anomalies(site_id, db)
    # Snapshot optionnel pour améliorer SURFACE_MISMATCH (usage-aware)
    snapshot = get_site_snapshot(site_id, org_id, db) or {}
    enriched = enrich_anomalies_with_impact(result["anomalies"], snapshot, DEFAULT_ASSUMPTIONS)
    total_risk_eur = sum((a.get("business_impact") or {}).get("estimated_risk_eur") or 0.0 for a in enriched)
    return {
        **result,
        "anomalies": enriched,
        "total_estimated_risk_eur": round(total_risk_eur, 0),
        "assumptions_used": DEFAULT_ASSUMPTIONS.to_dict(),
    }


@router.get("/anomalies", response_model=OrgAnomaliesResponse)
def list_org_anomalies(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    min_score: Optional[int] = Query(None, ge=0, le=100, description="Filtre sites avec score ≤ min_score"),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """
    Liste paginée des sites de l'org avec leurs anomalies patrimoine (V59).
    Chaque anomalie enrichie : regulatory_impact, business_impact, priority_score.
    Triée par completude_score ASC (plus dégradés en premier).
    """
    from services.patrimoine_anomalies import compute_site_anomalies
    from services.patrimoine_impact import enrich_anomalies_with_impact
    from config.patrimoine_assumptions import DEFAULT_ASSUMPTIONS

    org_id = _get_org_id(request, auth, db)

    sites_q = (
        db.query(Site)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(Site.actif.is_(True))
        .order_by(Site.id)
    )
    all_sites = sites_q.all()

    results = []
    for site in all_sites:
        data = compute_site_anomalies(site.id, db)
        if min_score is not None and data["completude_score"] > min_score:
            continue
        enriched = enrich_anomalies_with_impact(data["anomalies"], None, DEFAULT_ASSUMPTIONS)
        total_risk_eur = sum((a.get("business_impact") or {}).get("estimated_risk_eur") or 0.0 for a in enriched)
        top_priority = enriched[0]["priority_score"] if enriched else None
        results.append(
            {
                "site_id": site.id,
                "nom": site.nom,
                "completude_score": data["completude_score"],
                "nb_anomalies": data["nb_anomalies"],
                "top_severity": enriched[0]["severity"] if enriched else None,
                "top_priority_score": top_priority,
                "total_estimated_risk_eur": round(total_risk_eur, 0),
                "anomalies": enriched,
            }
        )

    # Tri : scores les plus bas en premier (les plus à risque)
    results.sort(key=lambda r: r["completude_score"])

    total = len(results)
    offset = (page - 1) * page_size
    page_items = results[offset : offset + page_size]

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "sites": page_items,
    }


@router.get("/assumptions")
def get_patrimoine_assumptions():
    """
    Retourne les hypothèses de calcul d'impact en lecture seule (V59).
    Permet au frontend d'afficher la transparence des estimations.
    """
    from config.patrimoine_assumptions import DEFAULT_ASSUMPTIONS

    return DEFAULT_ASSUMPTIONS.to_dict()


@router.get("/portfolio-summary", response_model=PortfolioSummaryResponse)
def get_portfolio_summary(
    request: Request,
    portefeuille_id: Optional[int] = Query(None, description="Filtre par portefeuille"),
    site_id: Optional[int] = Query(None, description="Filtre par site unique"),
    top_n: int = Query(default=3, ge=1, le=10, description="Nombre de top sites retournés"),
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_portfolio_optional_auth),
):
    """
    Agrégation portfolio patrimoine : risque global, framework breakdown, top sites (V60).

    - Multi-org safe : scoped via org_id + filtres optionnels portefeuille/site.
    - Zéro N+1 côté query SQL — enrichissement impact fait en mémoire via enrich_anomalies_with_impact().
    - Cas critique : org vide ou scope vide → tout à 0, listes vides, pas de crash.
    - top_n (1..10, défaut 3) : contrôle la taille de top_sites.
    - Gracieux : si org non résolue (no auth, no demo) → 200 empty (jamais de 401/403).
    """
    from services.patrimoine_anomalies import compute_site_anomalies
    from services.patrimoine_impact import enrich_anomalies_with_impact
    from config.patrimoine_assumptions import DEFAULT_ASSUMPTIONS
    from services.patrimoine_portfolio_cache import get_prev_snapshot, set_snapshot

    # Résolution org_id gracieuse : si non résolu → 200 vide (pas de 401/403)
    # Évite le bandeau d'erreur frontend quand l'auth n'est pas encore établie.
    try:
        org_id = _get_org_id(request, auth, db)
    except HTTPException:
        # Org non résolue : pas d'auth, pas de DemoState, pas d'org active en DB.
        # Retourner une réponse vide valide plutôt qu'une erreur 401/403.
        from datetime import datetime as _dt

        return {
            "scope": {"org_id": None, "portefeuille_id": portefeuille_id, "site_id": site_id},
            "total_estimated_risk_eur": 0.0,
            "sites_count": 0,
            "sites_at_risk": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "sites_health": {"healthy": 0, "warning": 0, "critical": 0, "healthy_pct": 0.0},
            "framework_breakdown": [],
            "top_sites": [],
            "trend": None,
            "computed_at": _dt.utcnow().isoformat() + "Z",
        }

    _SEV_ORDER = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1}

    # Build sites query — même chaîne de jointures que list_org_anomalies
    sites_q = (
        db.query(Site)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(Site.actif.is_(True))
        .order_by(Site.id)
    )
    if portefeuille_id is not None:
        sites_q = sites_q.filter(Site.portefeuille_id == portefeuille_id)
    if site_id is not None:
        sites_q = sites_q.filter(Site.id == site_id)

    all_sites = sites_q.all()

    _HEALTH_HEALTHY = 85
    _HEALTH_WARNING = 50

    # Scope vide → tout à 0
    if not all_sites:
        computed_at_empty = datetime.now(timezone.utc).isoformat() + "Z"
        # Trend V62 : scope vide → on ne met pas en cache (pas de data utile)
        empty_resp = {
            "scope": {"org_id": org_id, "portefeuille_id": portefeuille_id, "site_id": site_id},
            "total_estimated_risk_eur": 0.0,
            "sites_count": 0,
            "sites_at_risk": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "sites_health": {"healthy": 0, "warning": 0, "critical": 0, "healthy_pct": 0.0},
            "framework_breakdown": [],
            "top_sites": [],
            "trend": None,
            "computed_at": computed_at_empty,
        }
        return empty_resp

    # Agrégation
    total_risk = 0.0
    sites_at_risk: Dict[str, int] = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    sites_health: Dict[str, Any] = {"healthy": 0, "warning": 0, "critical": 0}
    framework_totals: Dict[str, Dict] = {}
    site_summaries = []

    for site in all_sites:
        data = compute_site_anomalies(site.id, db)
        enriched = enrich_anomalies_with_impact(data["anomalies"], None, DEFAULT_ASSUMPTIONS)

        site_risk = sum((a.get("business_impact") or {}).get("estimated_risk_eur") or 0.0 for a in enriched)
        total_risk += site_risk

        # Pire sévérité du site → bucket sites_at_risk
        if enriched:
            worst_sev = max(
                (a["severity"] for a in enriched),
                key=lambda s: _SEV_ORDER.get(s, 0),
            ).lower()
            if worst_sev in sites_at_risk:
                sites_at_risk[worst_sev] += 1

        # V61 — santé par score de complétude (data quality)
        score = data.get("completude_score", 0)
        if score >= _HEALTH_HEALTHY:
            sites_health["healthy"] += 1
        elif score >= _HEALTH_WARNING:
            sites_health["warning"] += 1
        else:
            sites_health["critical"] += 1

        # Breakdown par framework réglementaire
        for a in enriched:
            fw = (a.get("regulatory_impact") or {}).get("framework", "NONE")
            if fw == "NONE":
                continue
            risk_a = (a.get("business_impact") or {}).get("estimated_risk_eur") or 0.0
            if fw not in framework_totals:
                framework_totals[fw] = {"risk_eur": 0.0, "anomalies_count": 0}
            framework_totals[fw]["risk_eur"] += risk_a
            framework_totals[fw]["anomalies_count"] += 1

        # Framework dominant du site (depuis l'anomalie top priority)
        top_fw: Optional[str] = None
        if enriched:
            ri = enriched[0].get("regulatory_impact") or {}
            fw0 = ri.get("framework", "NONE")
            top_fw = fw0 if fw0 != "NONE" else None

        site_summaries.append(
            {
                "site_id": site.id,
                "site_nom": site.nom,
                "risk_eur": round(site_risk, 0),
                "anomalies_count": data["nb_anomalies"],
                "top_framework": top_fw,
            }
        )

    # healthy_pct final
    n_total = len(all_sites)
    sites_health["healthy_pct"] = round(sites_health["healthy"] / n_total * 100, 1) if n_total else 0.0

    # Top N sites par risk_eur DESC
    site_summaries.sort(key=lambda s: s["risk_eur"], reverse=True)
    top_sites = site_summaries[:top_n]

    # Framework breakdown trié par risk_eur DESC
    framework_breakdown = [
        {
            "framework": fw,
            "risk_eur": round(v["risk_eur"], 0),
            "anomalies_count": v["anomalies_count"],
        }
        for fw, v in sorted(framework_totals.items(), key=lambda x: x[1]["risk_eur"], reverse=True)
    ]

    computed_at = datetime.now(timezone.utc).isoformat() + "Z"
    total_risk_rounded = round(total_risk, 0)

    # V62 — Trend réel via snapshot in-memory par org_id
    # On ne cache que lorsque le scope est global (pas de filtre site/portefeuille)
    # pour éviter de polluer la baseline avec une vue partielle.
    _EPS = 1.0  # €  — seuil anti-bruit
    trend_payload: Optional[Dict[str, Any]] = None

    if portefeuille_id is None and site_id is None:
        prev = get_prev_snapshot(org_id)
        if prev is not None:
            delta_risk = total_risk_rounded - prev["total_estimated_risk_eur"]
            delta_sites = n_total - prev["sites_count"]
            if delta_risk > _EPS:
                direction = "up"
            elif delta_risk < -_EPS:
                direction = "down"
            else:
                direction = "stable"
            trend_payload = {
                "risk_eur_delta": round(delta_risk, 0),
                "sites_count_delta": delta_sites,
                "direction": direction,
                "vs_computed_at": prev["computed_at"],
            }
        # Mettre à jour le snapshot courant APRÈS avoir lu le précédent
        set_snapshot(
            org_id,
            {
                "computed_at": computed_at,
                "total_estimated_risk_eur": total_risk_rounded,
                "sites_count": n_total,
            },
        )

    return {
        "scope": {"org_id": org_id, "portefeuille_id": portefeuille_id, "site_id": site_id},
        "total_estimated_risk_eur": total_risk_rounded,
        "sites_count": n_total,
        "sites_at_risk": sites_at_risk,
        "sites_health": sites_health,
        "framework_breakdown": framework_breakdown,
        "top_sites": top_sites,
        "trend": trend_payload,
        "computed_at": computed_at,
    }


@router.get("/contracts")
def list_contracts(
    request: Request,
    site_id: Optional[int] = None,
    energy_type: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List energy contracts with filters — scoped to org."""
    org_id = _get_org_id(request, auth, db)
    q = (
        db.query(EnergyContract)
        .join(Site, EnergyContract.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
    )
    if site_id is not None:
        q = q.filter(EnergyContract.site_id == site_id)
    if energy_type:
        q = q.filter(EnergyContract.energy_type == energy_type)
    total = q.count()
    contracts = q.offset(skip).limit(limit).all()
    return {"total": total, "contracts": [_serialize_contract(ct) for ct in contracts]}


@router.post("/contracts")
def create_contract(
    request: Request,
    body: ContractCreateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Create a new energy contract."""
    org_id = _get_org_id(request, auth, db)
    site = _load_site_with_org_check(db, body.site_id, org_id)

    try:
        et = BillingEnergyType(body.energy_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Type energie invalide: {body.energy_type}")

    # V96 — parse optional enums
    offer_idx = None
    if body.offer_indexation:
        try:
            offer_idx = ContractIndexation(body.offer_indexation)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Indexation invalide: {body.offer_indexation}")
    ct_status = None
    if body.contract_status:
        try:
            ct_status = ContractStatus(body.contract_status)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Statut contrat invalide: {body.contract_status}")

    ct = EnergyContract(
        site_id=body.site_id,
        energy_type=et,
        supplier_name=body.supplier_name,
        start_date=date.fromisoformat(body.start_date) if body.start_date else None,
        end_date=date.fromisoformat(body.end_date) if body.end_date else None,
        price_ref_eur_per_kwh=body.price_ref_eur_per_kwh,
        fixed_fee_eur_per_month=body.fixed_fee_eur_per_month,
        notice_period_days=body.notice_period_days,
        auto_renew=body.auto_renew,
        offer_indexation=offer_idx,
        price_granularity=body.price_granularity,
        renewal_alert_days=body.renewal_alert_days,
        contract_status=ct_status,
    )
    db.add(ct)
    db.commit()
    db.refresh(ct)
    return _serialize_contract(ct)


@router.patch("/contracts/{contract_id}")
def update_contract(
    contract_id: int,
    request: Request,
    body: ContractUpdateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Update an energy contract (partial update)."""
    org_id = _get_org_id(request, auth, db)
    ct = _load_contract_with_org_check(db, contract_id, org_id)

    updates = body.model_dump(exclude_unset=True)

    # Apply field values (parse dates + V96 enums)
    for field, value in updates.items():
        if field in ("start_date", "end_date") and value is not None:
            value = date.fromisoformat(value)
        elif field == "offer_indexation" and value is not None:
            try:
                value = ContractIndexation(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Indexation invalide: {value}")
        elif field == "contract_status" and value is not None:
            try:
                value = ContractStatus(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Statut contrat invalide: {value}")
        setattr(ct, field, value)

    # If dates changed, check for overlap with other contracts
    if "start_date" in updates or "end_date" in updates:
        overlap = check_contract_overlap(
            db,
            ct.site_id,
            ct.energy_type,
            ct.start_date,
            ct.end_date,
            exclude_id=ct.id,
        )
        if overlap:
            raise HTTPException(
                status_code=409,
                detail=f"Chevauchement avec le contrat #{overlap.id} "
                f"({overlap.supplier_name}, "
                f"{overlap.start_date or '...'} → {overlap.end_date or '...'})",
            )

    db.commit()
    return {"updated": list(updates.keys()), **_serialize_contract(ct)}


@router.delete("/contracts/{contract_id}")
def delete_contract(
    contract_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Delete an energy contract."""
    org_id = _get_org_id(request, auth, db)
    ct = _load_contract_with_org_check(db, contract_id, org_id)
    db.delete(ct)
    db.commit()
    return {"detail": f"Contrat {contract_id} supprime"}


# ========================================
# V96: Payment Rules CRUD
# ========================================


def _serialize_payment_rule(pr: PaymentRule) -> dict:
    return {
        "id": pr.id,
        "level": pr.level.value if pr.level else None,
        "portefeuille_id": pr.portefeuille_id,
        "site_id": pr.site_id,
        "contract_id": pr.contract_id,
        "invoice_entity_id": pr.invoice_entity_id,
        "payer_entity_id": pr.payer_entity_id,
        "cost_center": pr.cost_center,
        "created_at": pr.created_at.isoformat() if pr.created_at else None,
    }


def _resolve_payment_rule(db: Session, site_id: int, contract_id: int = None) -> Optional[PaymentRule]:
    """Cascade resolution: contrat > site > portefeuille > None."""
    # 1. Contract-level
    if contract_id:
        pr = (
            db.query(PaymentRule)
            .filter(
                PaymentRule.level == PaymentRuleLevel.CONTRAT,
                PaymentRule.contract_id == contract_id,
            )
            .first()
        )
        if pr:
            return pr

    # 2. Site-level
    pr = (
        db.query(PaymentRule)
        .filter(
            PaymentRule.level == PaymentRuleLevel.SITE,
            PaymentRule.site_id == site_id,
        )
        .first()
    )
    if pr:
        return pr

    # 3. Portefeuille-level
    site = db.query(Site).filter(Site.id == site_id).first()
    if site and site.portefeuille_id:
        pr = (
            db.query(PaymentRule)
            .filter(
                PaymentRule.level == PaymentRuleLevel.PORTEFEUILLE,
                PaymentRule.portefeuille_id == site.portefeuille_id,
            )
            .first()
        )
        if pr:
            return pr

    return None


@router.get("/payment-rules")
def list_payment_rules(
    request: Request,
    level: Optional[str] = None,
    portefeuille_id: Optional[int] = None,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """List payment rules — scoped to org."""
    org_id = _get_org_id(request, auth, db)
    q = (
        db.query(PaymentRule)
        .join(EntiteJuridique, PaymentRule.invoice_entity_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
    )
    if level:
        try:
            q = q.filter(PaymentRule.level == PaymentRuleLevel(level))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Niveau invalide: {level}")
    if portefeuille_id is not None:
        q = q.filter(PaymentRule.portefeuille_id == portefeuille_id)
    rules = q.all()
    return {"rules": [_serialize_payment_rule(pr) for pr in rules]}


@router.post("/payment-rules")
def create_payment_rule(
    request: Request,
    body: PaymentRuleCreateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Create or upsert a payment rule at any level."""
    org_id = _get_org_id(request, auth, db)

    try:
        lvl = PaymentRuleLevel(body.level)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Niveau invalide: {body.level}")

    # Validate entity belongs to org
    ej = (
        db.query(EntiteJuridique)
        .filter(
            EntiteJuridique.id == body.invoice_entity_id,
            EntiteJuridique.organisation_id == org_id,
        )
        .first()
    )
    if not ej:
        raise HTTPException(status_code=404, detail="Entite juridique facturee non trouvee")

    if body.payer_entity_id:
        pej = (
            db.query(EntiteJuridique)
            .filter(
                EntiteJuridique.id == body.payer_entity_id,
                EntiteJuridique.organisation_id == org_id,
            )
            .first()
        )
        if not pej:
            raise HTTPException(status_code=404, detail="Entite juridique payeuse non trouvee")

    # Validate scope target
    if lvl == PaymentRuleLevel.PORTEFEUILLE and body.portefeuille_id:
        _check_portfolio_belongs_to_org(db, body.portefeuille_id, org_id)
    elif lvl == PaymentRuleLevel.SITE and body.site_id:
        _load_site_with_org_check(db, body.site_id, org_id)
    elif lvl == PaymentRuleLevel.CONTRAT and body.contract_id:
        _load_contract_with_org_check(db, body.contract_id, org_id)

    # Upsert: check for existing rule at same scope
    existing = (
        db.query(PaymentRule)
        .filter(
            PaymentRule.level == lvl,
            PaymentRule.portefeuille_id == body.portefeuille_id,
            PaymentRule.site_id == body.site_id,
            PaymentRule.contract_id == body.contract_id,
        )
        .first()
    )

    if existing:
        existing.invoice_entity_id = body.invoice_entity_id
        existing.payer_entity_id = body.payer_entity_id
        existing.cost_center = body.cost_center
        db.commit()
        return _serialize_payment_rule(existing)

    pr = PaymentRule(
        level=lvl,
        portefeuille_id=body.portefeuille_id,
        site_id=body.site_id,
        contract_id=body.contract_id,
        invoice_entity_id=body.invoice_entity_id,
        payer_entity_id=body.payer_entity_id,
        cost_center=body.cost_center,
    )
    db.add(pr)
    db.commit()
    db.refresh(pr)
    return _serialize_payment_rule(pr)


@router.put("/payment-rules/{rule_id}")
def update_payment_rule(
    rule_id: int,
    request: Request,
    body: PaymentRuleCreateRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Update an existing payment rule."""
    org_id = _get_org_id(request, auth, db)
    pr = db.query(PaymentRule).filter(PaymentRule.id == rule_id).first()
    if not pr:
        raise HTTPException(status_code=404, detail=f"Regle {rule_id} non trouvee")

    # Check org ownership via invoice entity
    ej = (
        db.query(EntiteJuridique)
        .filter(
            EntiteJuridique.id == pr.invoice_entity_id,
            EntiteJuridique.organisation_id == org_id,
        )
        .first()
    )
    if not ej:
        raise HTTPException(status_code=404, detail=f"Regle {rule_id} non trouvee")

    pr.invoice_entity_id = body.invoice_entity_id
    pr.payer_entity_id = body.payer_entity_id
    pr.cost_center = body.cost_center
    db.commit()
    return _serialize_payment_rule(pr)


@router.delete("/payment-rules/{rule_id}")
def delete_payment_rule(
    rule_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Delete a payment rule."""
    org_id = _get_org_id(request, auth, db)
    pr = db.query(PaymentRule).filter(PaymentRule.id == rule_id).first()
    if not pr:
        raise HTTPException(status_code=404, detail=f"Regle {rule_id} non trouvee")

    ej = (
        db.query(EntiteJuridique)
        .filter(
            EntiteJuridique.id == pr.invoice_entity_id,
            EntiteJuridique.organisation_id == org_id,
        )
        .first()
    )
    if not ej:
        raise HTTPException(status_code=404, detail=f"Regle {rule_id} non trouvee")

    db.delete(pr)
    db.commit()
    return {"detail": f"Regle {rule_id} supprimee"}


@router.post("/payment-rules/apply-bulk")
def apply_payment_rules_bulk(
    request: Request,
    body: PaymentRuleBulkApplyRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Apply payment rule to N sites atomically."""
    org_id = _get_org_id(request, auth, db)

    ej = (
        db.query(EntiteJuridique)
        .filter(
            EntiteJuridique.id == body.invoice_entity_id,
            EntiteJuridique.organisation_id == org_id,
        )
        .first()
    )
    if not ej:
        raise HTTPException(status_code=404, detail="Entite juridique non trouvee")

    db.begin_nested()  # SAVEPOINT
    created = 0
    for sid in body.site_ids:
        _load_site_with_org_check(db, sid, org_id)
        existing = (
            db.query(PaymentRule)
            .filter(
                PaymentRule.level == PaymentRuleLevel.SITE,
                PaymentRule.site_id == sid,
            )
            .first()
        )
        if existing:
            existing.invoice_entity_id = body.invoice_entity_id
            existing.payer_entity_id = body.payer_entity_id
            existing.cost_center = body.cost_center
        else:
            db.add(
                PaymentRule(
                    level=PaymentRuleLevel.SITE,
                    site_id=sid,
                    invoice_entity_id=body.invoice_entity_id,
                    payer_entity_id=body.payer_entity_id,
                    cost_center=body.cost_center,
                )
            )
            created += 1

    db.commit()
    return {"applied": len(body.site_ids), "created": created}


@router.get("/sites/{site_id}/payment-info")
def get_site_payment_info(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Resolve effective payment rule for a site (contrat > site > portefeuille)."""
    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)

    pr = _resolve_payment_rule(db, site_id)
    if not pr:
        return {"resolved": False, "rule": None, "source_level": None}

    # Load entity names
    inv_ej = db.query(EntiteJuridique).get(pr.invoice_entity_id)
    pay_ej = db.query(EntiteJuridique).get(pr.payer_entity_id) if pr.payer_entity_id else None

    return {
        "resolved": True,
        "source_level": pr.level.value if pr.level else None,
        "rule": _serialize_payment_rule(pr),
        "invoice_entity_name": inv_ej.nom if inv_ej else None,
        "payer_entity_name": pay_ej.nom if pay_ej else None,
    }


# ========================================
# V96: Reconciliation endpoints
# ========================================


@router.get("/sites/{site_id}/reconciliation")
def get_site_reconciliation(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """3-way reconciliation for a single site."""
    from services.reconciliation_service import reconcile_site

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)
    return reconcile_site(db, site_id)


@router.get("/portfolio/reconciliation")
def get_portfolio_reconciliation(
    request: Request,
    portefeuille_id: Optional[int] = None,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """Aggregate reconciliation across all sites in scope."""
    from services.reconciliation_service import reconcile_portfolio

    org_id = _get_org_id(request, auth, db)
    if portefeuille_id:
        _check_portfolio_belongs_to_org(db, portefeuille_id, org_id)
    return reconcile_portfolio(db, org_id, portefeuille_id)


# ========================================
# V97: Resolution Engine endpoints
# ========================================


@router.post("/sites/{site_id}/reconciliation/fix")
def apply_reconciliation_fix(
    site_id: int,
    request: Request,
    body: ReconciliationFixRequest,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V97: Apply a 1-click fix for a reconciliation check."""
    from services.reconciliation_service import (
        fix_create_delivery_point,
        fix_extend_contract,
        fix_adjust_contract_dates,
        fix_align_energy_type,
        fix_create_payment_rule,
    )

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)

    params = body.params or {}
    applied_by = auth.user_email if auth and hasattr(auth, "user_email") else None

    FIXERS = {
        "create_delivery_point": fix_create_delivery_point,
        "extend_contract": fix_extend_contract,
        "adjust_contract_dates": fix_adjust_contract_dates,
        "align_energy_type": fix_align_energy_type,
        "create_payment_rule": fix_create_payment_rule,
    }

    fixer = FIXERS.get(body.action)
    if not fixer:
        raise HTTPException(status_code=400, detail=f"Action inconnue: {body.action}")

    db.begin_nested()
    result = fixer(db, site_id, **params, applied_by=applied_by)
    db.commit()

    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])

    return {"ok": True, "action": body.action, "result": result}


@router.get("/sites/{site_id}/reconciliation/history")
def get_reconciliation_fix_history(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V97: Get audit trail for reconciliation fixes on a site."""
    from services.reconciliation_service import get_fix_logs

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)
    return {"site_id": site_id, "logs": get_fix_logs(db, site_id)}


@router.get("/sites/{site_id}/reconciliation/evidence")
def get_reconciliation_evidence(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V97 Phase 4: Get evidence pack (JSON) for a site's reconciliation."""
    from services.reconciliation_service import get_evidence_pack

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)
    return get_evidence_pack(db, site_id)


@router.get("/sites/{site_id}/reconciliation/evidence/summary")
def get_reconciliation_evidence_summary(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V98: Get 1-page evidence summary for a site's reconciliation."""
    from services.reconciliation_service import get_evidence_summary

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)
    return get_evidence_summary(db, site_id)


@router.get("/sites/{site_id}/reconciliation/evidence/csv")
def get_reconciliation_evidence_csv(
    site_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V97 Phase 4: Export evidence pack as CSV."""
    from services.reconciliation_service import get_evidence_pack

    org_id = _get_org_id(request, auth, db)
    _load_site_with_org_check(db, site_id, org_id)
    pack = get_evidence_pack(db, site_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["check_id", "label_fr", "status", "reason_fr", "suggestion_fr"])
    for check in pack["reconciliation"]["checks"]:
        writer.writerow(
            [
                check["id"],
                check["label_fr"],
                check["status"],
                check["reason_fr"],
                check.get("suggestion_fr", ""),
            ]
        )
    writer.writerow([])
    writer.writerow(["fix_id", "check_id", "action", "status_before", "status_after", "applied_by", "applied_at"])
    for log in pack["fix_history"]:
        writer.writerow(
            [
                log["id"],
                log["check_id"],
                log["action"],
                log["status_before"],
                log["status_after"],
                log.get("applied_by", ""),
                log.get("applied_at", ""),
            ]
        )

    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=evidence_site_{site_id}.csv"},
    )


@router.get("/portfolio/reconciliation/evidence/csv")
def get_portfolio_evidence_csv(
    request: Request,
    portefeuille_id: Optional[int] = None,
    db: Session = Depends(get_db),
    auth: Optional[AuthContext] = Depends(get_optional_auth),
):
    """V97 Phase 4: Export portfolio reconciliation summary as CSV."""
    from services.reconciliation_service import reconcile_portfolio

    org_id = _get_org_id(request, auth, db)
    if portefeuille_id:
        _check_portfolio_belongs_to_org(db, portefeuille_id, org_id)
    data = reconcile_portfolio(db, org_id, portefeuille_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["site_id", "nom", "status", "score"])
    for s in data["sites"]:
        writer.writerow([s["site_id"], s["nom"], s["status"], s["score"]])
    writer.writerow([])
    writer.writerow(["stat", "value"])
    for k, v in data["stats"].items():
        writer.writerow([k, v])

    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=portfolio_reconciliation.csv"},
    )
