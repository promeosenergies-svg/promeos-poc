"""
PROMEOS - Patrimoine helpers partagés entre sous-modules.
Scope org, vérifications batch/site/compteur/contrat, sérialiseurs.
"""

import csv
import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, Optional, List

from fastapi import HTTPException, Request
from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from database import get_db
from models import (
    Organisation,
    EntiteJuridique,
    Portefeuille,
    StagingBatch,
    StagingSite,
    StagingCompteur,
    QualityFinding,
    ImportSourceType,
    StagingStatus,
    QualityRuleSeverity,
    ActivationLog,
    ActivationLogStatus,
    Site,
    DeliveryPoint,
    not_deleted,
    Compteur,
    TypeSite,
    TypeCompteur,
    EnergyVector,
    EnergyContract,
    BillingEnergyType,
    StatutConformite,
    PaymentRule,
    PaymentRuleLevel,
    ContractIndexation,
    ContractStatus,
    Batiment,
)
from middleware.auth import get_optional_auth, get_portfolio_optional_auth, AuthContext


# ========================================
# Multi-org scope helpers
# ========================================


def _get_org_id(request: Request, auth: Optional[AuthContext], db: Session) -> int:
    """Resolve org_id via centralized scope chain (DEMO_MODE-aware)."""
    from services.scope_utils import resolve_org_id

    return resolve_org_id(request, auth, db)


def _check_batch_org(batch: StagingBatch, org_id: int):
    """Verify batch belongs to the resolved org. Raises 403 if mismatch."""
    if batch is None:
        raise HTTPException(status_code=404, detail="Batch non trouvé")
    if batch.org_id is not None and batch.org_id != org_id:
        raise HTTPException(status_code=403, detail="Batch hors périmètre")


def _check_site_belongs_to_org(db: Session, site: Site, org_id: int):
    """Verify site belongs to org via portfolio→EJ chain. Fail-closed: raises 403 on any break."""
    if not site.portefeuille_id:
        raise HTTPException(status_code=403, detail="Site hors périmètre")
    pf = db.get(Portefeuille, site.portefeuille_id)
    if not pf:
        raise HTTPException(status_code=403, detail="Site hors périmètre")
    ej = db.get(EntiteJuridique, pf.entite_juridique_id)
    if not ej or ej.organisation_id != org_id:
        raise HTTPException(status_code=403, detail="Site hors périmètre")


def _check_portfolio_belongs_to_org(db: Session, portfolio_id: int, org_id: int):
    """Verify portfolio belongs to org. Raises 403 if mismatch."""
    pf = db.get(Portefeuille, portfolio_id)
    if not pf:
        raise HTTPException(status_code=404, detail=f"Portefeuille {portfolio_id} non trouvé")
    ej = db.get(EntiteJuridique, pf.entite_juridique_id)
    if not ej or ej.organisation_id != org_id:
        raise HTTPException(status_code=403, detail="Portefeuille hors périmètre")
    return pf


def _load_site_with_org_check(db: Session, site_id: int, org_id: int) -> Site:
    """Load a site and verify org ownership. Raises 404/403."""
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail=f"Site {site_id} non trouvé")
    _check_site_belongs_to_org(db, site, org_id)
    return site


def _load_compteur_with_org_check(db: Session, compteur_id: int, org_id: int) -> "Compteur":
    """Load a compteur with upfront org verification via JOIN chain. Returns 404 on miss."""
    c = (
        db.query(Compteur)
        .join(Site, Compteur.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(Compteur.id == compteur_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail=f"Compteur {compteur_id} non trouvé")
    return c


def _load_contract_with_org_check(db: Session, contract_id: int, org_id: int) -> "EnergyContract":
    """Load a contract with upfront org verification via JOIN chain. Returns 404 on miss."""
    ct = (
        db.query(EnergyContract)
        .join(Site, EnergyContract.site_id == Site.id)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id)
        .filter(EnergyContract.id == contract_id)
        .first()
    )
    if not ct:
        raise HTTPException(status_code=404, detail=f"Contrat {contract_id} non trouvé")
    return ct


# ========================================
# Helpers
# ========================================


def _normalize_compteur_type(raw: str) -> str:
    """Normalize compteur type string for autofix."""
    low = raw.lower().strip()
    if any(k in low for k in ("elec", "elect")):
        return "electricite"
    if any(k in low for k in ("gaz", "gas")):
        return "gaz"
    if any(k in low for k in ("eau", "water")):
        return "eau"
    return raw


def _parse_excel_to_staging(db: Session, batch_id: int, content: bytes) -> dict:
    """Parse Excel file via openpyxl and feed into staging."""
    import io as _io
    from openpyxl import load_workbook
    from services.import_mapping import normalize_column_name
    from services.patrimoine_service import import_csv_to_staging

    wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"sites_count": 0, "compteurs_count": 0, "parse_errors": [{"row": 0, "error": "Empty workbook"}]}

    # First row = headers — normalize via mapping
    raw_headers = [str(h or "").strip() for h in rows[0]]
    normalized_headers = [normalize_column_name(h) for h in raw_headers]

    # Convert to CSV bytes with normalized headers
    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(normalized_headers)
    for row in rows[1:]:
        writer.writerow([str(c) if c is not None else "" for c in row])

    csv_bytes = output.getvalue().encode("utf-8")
    return import_csv_to_staging(db, batch_id, csv_bytes)


# ========================================
# Serializers
# ========================================


def _worst_compliance_status(*statuses) -> StatutConformite | None:
    """Return the most non-compliant status across all given frameworks.

    Priority: NON_CONFORME > A_RISQUE > CONFORME.
    None values are ignored (framework not applicable).
    """
    valid = [s for s in statuses if s is not None]
    if not valid:
        return None
    if any(s == StatutConformite.NON_CONFORME for s in valid):
        return StatutConformite.NON_CONFORME
    if any(s == StatutConformite.A_RISQUE for s in valid):
        return StatutConformite.A_RISQUE
    return StatutConformite.CONFORME


def _compute_compliance_review_status(site, db=None) -> dict:
    """Compute whether this site needs compliance review and why."""
    reasons = []

    # 1. Statut non conforme
    dt = str(getattr(site, "statut_decret_tertiaire", ""))
    bacs = str(getattr(site, "statut_bacs", ""))
    if "NON_CONFORME" in dt or "NON_CONFORME" in bacs:
        reasons.append("non_conforme")
    elif "A_RISQUE" in dt or "A_RISQUE" in bacs:
        reasons.append("a_risque")

    # 2. Incomplete patrimoine
    if not site.surface_m2 or site.surface_m2 <= 0:
        reasons.append("surface_manquante")
    if not site.siret:
        reasons.append("siret_manquant")

    # 3. High financial risk
    risk = getattr(site, "risque_financier_euro", 0) or 0
    if risk > 10000:
        reasons.append("risque_eleve")

    # 4. Score too low
    score = getattr(site, "compliance_score_composite", None)
    if score is not None and score < 50:
        reasons.append("score_critique")

    return {
        "needs_review": len(reasons) > 0,
        "reasons": reasons,
        "reason_count": len(reasons),
    }


def _serialize_site(site: Site) -> dict:
    review_status = _compute_compliance_review_status(site)
    return {
        "id": site.id,
        "nom": site.nom,
        "type": site.type.value if site.type else None,
        "usage": site.type.value if site.type else None,
        "adresse": site.adresse,
        "code_postal": site.code_postal,
        "ville": site.ville,
        "region": site.region,
        "surface_m2": site.surface_m2,
        "nombre_employes": site.nombre_employes,
        "siret": site.siret,
        "naf_code": site.naf_code,
        "actif": site.actif,
        "portefeuille_id": site.portefeuille_id,
        "portefeuille_nom": site.portefeuille.nom if site.portefeuille else None,
        "data_source": site.data_source,
        "created_at": site.created_at.isoformat() if site.created_at else None,
        "updated_at": site.updated_at.isoformat() if site.updated_at else None,
        # Enriched analytics fields
        "risque_eur": site.risque_financier_euro,
        "statut_conformite": (
            _worst_compliance_status(site.statut_decret_tertiaire, site.statut_bacs).value
            if _worst_compliance_status(site.statut_decret_tertiaire, site.statut_bacs)
            else None
        ),
        "anomalie_facture": site.anomalie_facture,
        "conso_kwh_an": site.annual_kwh_total,
        "compliance_needs_review": review_status["needs_review"],
        "compliance_review_reasons": review_status["reasons"],
    }


def _build_sites_query(
    db: Session, org_id: int, portefeuille_id=None, actif=None, ville=None, type_site=None, search=None
):
    """Build a filtered site query scoped to org — shared by list_sites and export."""
    q = (
        db.query(Site)
        .join(Portefeuille, Site.portefeuille_id == Portefeuille.id)
        .join(EntiteJuridique, Portefeuille.entite_juridique_id == EntiteJuridique.id)
        .filter(EntiteJuridique.organisation_id == org_id, not_deleted(Site))
    )
    if portefeuille_id is not None:
        q = q.filter(Site.portefeuille_id == portefeuille_id)
    if actif is not None:
        q = q.filter(Site.actif == actif)
    if ville:
        q = q.filter(Site.ville.ilike(f"%{ville}%"))
    if type_site:
        q = q.filter(Site.type == type_site)
    if search:
        q = q.filter(
            (Site.nom.ilike(f"%{search}%")) | (Site.ville.ilike(f"%{search}%")) | (Site.adresse.ilike(f"%{search}%"))
        )
    return q


def _serialize_compteur(c: Compteur) -> dict:
    return {
        "id": c.id,
        "source": "compteur",
        "site_id": c.site_id,
        "type": c.type.value if c.type else None,
        "numero_serie": c.numero_serie,
        "meter_id": c.meter_id,
        "puissance_souscrite_kw": c.puissance_souscrite_kw,
        "energy_vector": c.energy_vector.value if c.energy_vector else None,
        "actif": c.actif,
        "data_source": c.data_source,
    }


def _serialize_contract(ct: EnergyContract) -> dict:
    return {
        "id": ct.id,
        "site_id": ct.site_id,
        "energy_type": ct.energy_type.value if ct.energy_type else None,
        "supplier_name": ct.supplier_name,
        "start_date": ct.start_date.isoformat() if ct.start_date else None,
        "end_date": ct.end_date.isoformat() if ct.end_date else None,
        "price_ref_eur_per_kwh": ct.price_ref_eur_per_kwh,
        "fixed_fee_eur_per_month": ct.fixed_fee_eur_per_month,
        "notice_period_days": ct.notice_period_days,
        "auto_renew": ct.auto_renew,
        # V96
        "offer_indexation": ct.offer_indexation.value if ct.offer_indexation else None,
        "price_granularity": ct.price_granularity,
        "renewal_alert_days": ct.renewal_alert_days,
        "contract_status": ct.contract_status.value if ct.contract_status else None,
        "created_at": ct.created_at.isoformat() if ct.created_at else None,
        # V-registre: champs registre patrimonial & contractuel
        "reference_fournisseur": ct.reference_fournisseur,
        "date_signature": ct.date_signature.isoformat() if ct.date_signature else None,
        "conditions_particulieres": ct.conditions_particulieres,
        "document_url": ct.document_url,
        "delivery_point_ids": [dp.id for dp in ct.delivery_points] if ct.delivery_points else [],
        "delivery_points_count": len(ct.delivery_points) if ct.delivery_points else 0,
    }


def _serialize_payment_rule(pr: PaymentRule) -> dict:
    return {
        "id": pr.id,
        "level": pr.level.value if pr.level else None,
        "portefeuille_id": pr.portefeuille_id,
        "site_id": pr.site_id,
        "contract_id": pr.contract_id,
        "invoice_entity_id": pr.invoice_entity_id,
        "payer_entity_id": pr.payer_entity_id,
        "cost_center": pr.cost_center,
        "created_at": pr.created_at.isoformat() if pr.created_at else None,
    }


def _resolve_payment_rule(db: Session, site_id: int, contract_id: int = None) -> Optional[PaymentRule]:
    """Cascade resolution: contrat > site > portefeuille > None."""
    # 1. Contract-level
    if contract_id:
        pr = (
            db.query(PaymentRule)
            .filter(
                PaymentRule.level == PaymentRuleLevel.CONTRAT,
                PaymentRule.contract_id == contract_id,
            )
            .first()
        )
        if pr:
            return pr

    # 2. Site-level
    pr = (
        db.query(PaymentRule)
        .filter(
            PaymentRule.level == PaymentRuleLevel.SITE,
            PaymentRule.site_id == site_id,
        )
        .first()
    )
    if pr:
        return pr

    # 3. Portefeuille-level
    site = db.query(Site).filter(Site.id == site_id).first()
    if site and site.portefeuille_id:
        pr = (
            db.query(PaymentRule)
            .filter(
                PaymentRule.level == PaymentRuleLevel.PORTEFEUILLE,
                PaymentRule.portefeuille_id == site.portefeuille_id,
            )
            .first()
        )
        if pr:
            return pr

    return None


def _compute_site_completeness(db: Session, site, site_ids: list) -> dict:
    """Score de completude d'un site (0-100). Verifie les champs critiques du registre."""
    checks = {}
    # 1. Adresse
    checks["adresse"] = bool(site.adresse and site.ville)
    # 2. Surface
    checks["surface"] = bool(site.surface_m2 and site.surface_m2 > 0)
    # 3. Type site
    checks["type_site"] = bool(site.type)
    # 4. Entite juridique (via portefeuille)
    checks["entite_juridique"] = bool(site.portefeuille_id)
    # 5. Delivery point
    nb_dp = (
        db.query(DeliveryPoint)
        .filter(
            DeliveryPoint.site_id == site.id,
            not_deleted(DeliveryPoint),
        )
        .count()
    )
    checks["delivery_point"] = nb_dp > 0
    # 6. Contrat energie actif (end_date NULL = contrat en cours)
    today = date.today()
    nb_ct = (
        db.query(EnergyContract)
        .filter(
            EnergyContract.site_id == site.id,
            (EnergyContract.end_date >= today) | (EnergyContract.end_date.is_(None)),
        )
        .count()
    )
    checks["contrat_actif"] = nb_ct > 0
    # 7. Coordonnees GPS
    checks["coordonnees"] = bool(site.latitude and site.longitude)
    # 8. SIRET site
    checks["siret"] = bool(site.siret)

    filled = sum(1 for v in checks.values() if v)
    total = len(checks)
    score = round(filled / total * 100) if total else 0
    level = "complet" if score >= 80 else "partiel" if score >= 50 else "critique"
    missing = [k for k, v in checks.items() if not v]

    return {"score": score, "level": level, "filled": filled, "total": total, "missing": missing, "checks": checks}


# ========================================
# Pydantic Schemas (shared across modules)
# ========================================


class FixRequest(BaseModel):
    fix_type: str
    params: dict


class BulkFixRequest(BaseModel):
    fixes: List[FixRequest]


class ActivateRequest(BaseModel):
    portefeuille_id: int


class InvoiceImportRequest(BaseModel):
    invoices: list


class UpdateFieldRequest(BaseModel):
    staging_site_id: Optional[int] = None
    staging_compteur_id: Optional[int] = None
    field: str
    value: Optional[str] = None


class SiteUpdateRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    nom: Optional[str] = None
    adresse: Optional[str] = None
    code_postal: Optional[str] = None
    ville: Optional[str] = None
    region: Optional[str] = None
    surface_m2: Optional[float] = Field(None, gt=0, le=1_000_000)
    tertiaire_area_m2: Optional[float] = Field(None, ge=0)
    roof_area_m2: Optional[float] = Field(None, ge=0)
    parking_area_m2: Optional[float] = Field(None, ge=0)
    parking_type: Optional[str] = None
    nombre_employes: Optional[int] = Field(None, ge=0)
    naf_code: Optional[str] = None
    siret: Optional[str] = None
    type: Optional[str] = None
    latitude: Optional[float] = Field(None, ge=-90, le=90)
    longitude: Optional[float] = Field(None, ge=-180, le=180)
    is_multi_occupied: Optional[bool] = None

    @field_validator("code_postal")
    @classmethod
    def validate_cp(cls, v):
        import re

        if v is not None and not re.match(r"^\d{5}$", v):
            raise ValueError("Code postal invalide (5 chiffres attendus)")
        return v

    @field_validator("siret")
    @classmethod
    def validate_siret(cls, v):
        import re

        if v is not None and not re.match(r"^\d{14}$", v):
            raise ValueError("SIRET invalide (14 chiffres attendus)")
        return v

    @model_validator(mode="after")
    def check_surfaces(self):
        if self.tertiaire_area_m2 is not None and self.surface_m2 is not None:
            if self.tertiaire_area_m2 > self.surface_m2:
                raise ValueError("tertiaire_area_m2 ne peut pas dépasser surface_m2")
        return self


class SiteMergeRequest(BaseModel):
    source_site_id: int
    target_site_id: int


class CompteurMoveRequest(BaseModel):
    target_site_id: int


class CompteurUpdateRequest(BaseModel):
    numero_serie: Optional[str] = None
    meter_id: Optional[str] = None
    puissance_souscrite_kw: Optional[float] = None
    type: Optional[str] = None


class ContractCreateRequest(BaseModel):
    site_id: int = Field(..., gt=0, description="ID du site rattache")
    energy_type: str = Field("elec", description="Type energie: elec, gaz, eau, fioul, chaleur")
    supplier_name: str = Field(..., min_length=1, max_length=300, description="Nom du fournisseur")
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    price_ref_eur_per_kwh: Optional[float] = Field(None, ge=0, le=10)
    fixed_fee_eur_per_month: Optional[float] = Field(None, ge=0, le=1e6)
    notice_period_days: int = Field(90, ge=0, le=365)
    auto_renew: bool = False
    # V96
    offer_indexation: Optional[str] = None
    price_granularity: Optional[str] = None
    renewal_alert_days: Optional[int] = Field(None, ge=0, le=365)
    contract_status: Optional[str] = None
    # V-registre: champs registre patrimonial & contractuel
    reference_fournisseur: Optional[str] = Field(None, max_length=200)
    date_signature: Optional[str] = None
    conditions_particulieres: Optional[str] = Field(None, max_length=2000)
    document_url: Optional[str] = Field(None, max_length=500)
    delivery_point_ids: Optional[list] = None  # IDs des DP couverts


class ContractUpdateRequest(BaseModel):
    supplier_name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    price_ref_eur_per_kwh: Optional[float] = None
    fixed_fee_eur_per_month: Optional[float] = None
    notice_period_days: Optional[int] = None
    auto_renew: Optional[bool] = None
    # V96
    offer_indexation: Optional[str] = None
    price_granularity: Optional[str] = None
    renewal_alert_days: Optional[int] = None
    contract_status: Optional[str] = None
    # V-registre
    reference_fournisseur: Optional[str] = None
    date_signature: Optional[str] = None
    conditions_particulieres: Optional[str] = None
    document_url: Optional[str] = None
    delivery_point_ids: Optional[list] = None


# V96: Payment Rules schemas
class PaymentRuleCreateRequest(BaseModel):
    level: str  # portefeuille | site | contrat
    portefeuille_id: Optional[int] = None
    site_id: Optional[int] = None
    contract_id: Optional[int] = None
    invoice_entity_id: int
    payer_entity_id: Optional[int] = None
    cost_center: Optional[str] = None


class PaymentRuleBulkApplyRequest(BaseModel):
    site_ids: List[int]
    invoice_entity_id: int
    payer_entity_id: Optional[int] = None
    cost_center: Optional[str] = None


# V97: Resolution Engine schemas
class ReconciliationFixRequest(BaseModel):
    action: str
    params: Optional[Dict[str, Any]] = None


class MappingPreviewRequest(BaseModel):
    headers: list


class SubMeterCreateRequest(BaseModel):
    meter_id: Optional[str] = None
    name: Optional[str] = None
    numero_serie: Optional[str] = None
    type_compteur: Optional[str] = None
    subscribed_power_kva: Optional[float] = None


# ── Response models — KPIs & DeliveryPoints (V110) ──


class PatrimoineKpisResponse(BaseModel):
    total: int
    conformes: int
    aRisque: int
    nonConformes: int
    totalRisque: float
    totalSurface: float
    totalAnomalies: int
    nb_organisations: int
    nb_entites_juridiques: int
    nb_portefeuilles: int
    nb_sites: int
    nb_batiments: int
    nb_delivery_points: int
    nb_contrats: int
    nb_contrats_actifs: int
    nb_contrats_expiring_90j: int
    surface_totale_m2: float
    completude_moyenne_pct: int


class DeliveryPointItemResponse(BaseModel):
    id: int
    code: Optional[str] = None
    energy_type: Optional[str] = None
    status: Optional[str] = None
    compteurs_count: int = 0
    data_source: Optional[str] = None
    created_at: Optional[str] = None


class CompletenessResponse(BaseModel):
    score: int
    level: str
    filled: int
    total: int
    missing: List[str]
    checks: Dict[str, Any]


# ── Response models — Snapshot & Anomalies (V59) ──


class RegulatoryImpact(BaseModel):
    framework: str  # DECRET_TERTIAIRE / FACTURATION / BACS / NONE
    risk_level: str  # HIGH / MEDIUM / LOW
    explanation_fr: str


class BusinessImpact(BaseModel):
    type: str  # DATA_QUALITY / REGULATORY_RISK / BILLING_RISK
    estimated_risk_eur: float
    confidence: float  # 0..1
    explanation_fr: str


class AnomalyResponse(BaseModel):
    code: str
    severity: str
    title_fr: str
    detail_fr: str
    evidence: Dict[str, Any]
    cta: Dict[str, str]
    fix_hint_fr: str
    # V59 additions (always present, null-safe)
    regulatory_impact: Optional[RegulatoryImpact] = None
    business_impact: Optional[BusinessImpact] = None
    priority_score: Optional[int] = None


class SiteAnomaliesResponse(BaseModel):
    site_id: int
    anomalies: List[AnomalyResponse]
    completude_score: int
    nb_anomalies: int
    computed_at: str
    # V59 additions
    total_estimated_risk_eur: float
    assumptions_used: Dict[str, Any]


class UnifiedAnomalyItem(BaseModel):
    """Anomalie unifiée (patrimoine ou analytique KB)."""

    source: str  # "patrimoine" | "analytique"
    code: str
    severity: str  # "critical" | "high" | "medium" | "low"
    title_fr: str
    detail_fr: Optional[str] = None
    evidence: Optional[Dict[str, Any]] = None
    # Patrimoine-specific
    cta: Optional[Dict[str, str]] = None
    fix_hint_fr: Optional[str] = None
    regulatory_impact: Optional[RegulatoryImpact] = None
    business_impact: Optional[BusinessImpact] = None
    priority_score: Optional[int] = None
    # Analytique KB-specific
    confidence: Optional[float] = None
    deviation_pct: Optional[float] = None
    measured_value: Optional[float] = None
    threshold_value: Optional[float] = None


class UnifiedAnomaliesResponse(BaseModel):
    """Réponse endpoint anomalies-unified : patrimoine + KB combinés."""

    site_id: int
    anomalies: List[UnifiedAnomalyItem]
    total: int
    patrimoine_count: int
    analytique_count: int
    completude_score: Optional[int] = None
    computed_at: str


class OrgAnomaliesSiteItem(BaseModel):
    site_id: int
    nom: str
    completude_score: int
    nb_anomalies: int
    top_severity: Optional[str]
    top_priority_score: Optional[int]
    total_estimated_risk_eur: float
    anomalies: List[AnomalyResponse]


class OrgAnomaliesResponse(BaseModel):
    total: int
    page: int
    page_size: int
    sites: List[OrgAnomaliesSiteItem]


# ── V60/V61 : Portfolio summary ──


class PortfolioSitesAtRisk(BaseModel):
    critical: int = 0
    high: int = 0
    medium: int = 0
    low: int = 0


class PortfolioSitesHealth(BaseModel):
    """V61 — distribution des sites par score de complétude (data quality)."""

    healthy: int = 0  # completude_score >= 85
    warning: int = 0  # 50 <= completude_score < 85
    critical: int = 0  # completude_score < 50
    healthy_pct: float = 0.0


class PortfolioTrend(BaseModel):
    """V61 — tendance vs snapshot précédent. Null si pas d'historique."""

    risk_eur_delta: Optional[float] = None
    sites_count_delta: Optional[int] = None
    direction: Optional[str] = None  # "up" | "down" | "stable" | null
    vs_computed_at: Optional[str] = None


class PortfolioFrameworkItem(BaseModel):
    framework: str
    risk_eur: float
    anomalies_count: int


class PortfolioTopSiteItem(BaseModel):
    site_id: int
    site_nom: str
    risk_eur: float
    anomalies_count: int
    top_framework: Optional[str] = None


class PortfolioSummaryResponse(BaseModel):
    scope: Dict[str, Any]
    total_estimated_risk_eur: float
    sites_count: int
    sites_at_risk: PortfolioSitesAtRisk
    sites_health: PortfolioSitesHealth  # V61 NEW
    framework_breakdown: List[PortfolioFrameworkItem]
    top_sites: List[PortfolioTopSiteItem]
    trend: Optional[PortfolioTrend] = None  # V61 NEW (null — pas d'historique encore)
    computed_at: str


# ── Response models — Meters unified (V110) ──


class SubMeterItemResponse(BaseModel):
    id: int
    meter_id: Optional[str] = None
    name: Optional[str] = None
    energy_vector: Optional[str] = None


class MeterItemResponse(BaseModel):
    id: int
    source: str = "meter"
    meter_id: Optional[str] = None
    numero_serie: Optional[str] = None
    energy_vector: Optional[str] = None
    type_compteur: Optional[str] = None
    subscribed_power_kva: Optional[float] = None
    site_id: Optional[int] = None
    parent_meter_id: Optional[int] = None
    delivery_point_id: Optional[int] = None
    has_readings: bool = False
    is_active: bool = True
    name: Optional[str] = None
    marque: Optional[str] = None
    modele: Optional[str] = None
    sub_meters: List[SubMeterItemResponse] = []


class SiteMetersResponse(BaseModel):
    meters: List[MeterItemResponse]
