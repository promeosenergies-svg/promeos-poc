"""
PROMEOS - Tests VNext: import_mapping + template endpoints + autofix + staging API.
Covers: column synonyms, normalization, delimiter/encoding detection, template download,
rows/issues/bulk-fix/autofix/result/export API endpoints.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import io
import csv
import json
import importlib
import pytest

_has_openpyxl = importlib.util.find_spec("openpyxl") is not None
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from models import (
    Base,
    Site,
    Compteur,
    Organisation,
    EntiteJuridique,
    Portefeuille,
    StagingBatch,
    StagingSite,
    StagingCompteur,
    QualityFinding,
    StagingStatus,
    ImportSourceType,
    QualityRuleSeverity,
    TypeSite,
    TypeCompteur,
    DeliveryPoint,
    DeliveryPointEnergyType,
    not_deleted,
)
from database import get_db
from main import app
from services.patrimoine_service import (
    create_staging_batch,
    import_csv_to_staging,
    get_staging_summary,
    run_quality_gate,
    activate_batch,
)
from services.import_mapping import (
    CANONICAL_COLUMNS,
    normalize_column_name,
    map_headers,
    detect_delimiter,
    detect_encoding,
    normalize_rows,
    generate_csv_template,
    generate_xlsx_template,
)


# ========================================
# Fixtures
# ========================================


@pytest.fixture
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        echo=False,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


@pytest.fixture
def client(db_session):
    def _override():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = _override
    yield TestClient(app)
    app.dependency_overrides.clear()


def _create_org(db_session):
    org = Organisation(nom="Test Org", type_client="bureau", actif=True, siren="123456789")
    db_session.add(org)
    db_session.flush()
    ej = EntiteJuridique(organisation_id=org.id, nom="Test EJ", siren="123456789")
    db_session.add(ej)
    db_session.flush()
    pf = Portefeuille(entite_juridique_id=ej.id, nom="PF Test", description="Test")
    db_session.add(pf)
    db_session.flush()
    return org, ej, pf


def _create_batch_with_data(db_session, org_id=None):
    batch = create_staging_batch(
        db_session,
        org_id=org_id,
        user_id=None,
        source_type=ImportSourceType.CSV,
        mode="import",
    )
    s1 = StagingSite(
        batch_id=batch.id,
        row_number=2,
        nom="Site Alpha",
        adresse="10 rue de la Paix",
        code_postal="75001",
        ville="Paris",
        surface_m2=1200,
    )
    s2 = StagingSite(
        batch_id=batch.id,
        row_number=3,
        nom="Site Beta",
        adresse="20 avenue des Champs",
        code_postal="75008",
        ville="Paris",
        surface_m2=800,
    )
    db_session.add_all([s1, s2])
    db_session.flush()

    c1 = StagingCompteur(
        batch_id=batch.id,
        staging_site_id=s1.id,
        numero_serie="PRM-001",
        meter_id="12345678901234",
        type_compteur="electricite",
        puissance_kw=60,
    )
    c2 = StagingCompteur(batch_id=batch.id, staging_site_id=s1.id, numero_serie="PRM-002", type_compteur="gaz")
    c3 = StagingCompteur(
        batch_id=batch.id, staging_site_id=s2.id, numero_serie="PRM-003", type_compteur="electricite", puissance_kw=36
    )
    db_session.add_all([c1, c2, c3])
    db_session.flush()
    return batch, [s1, s2], [c1, c2, c3]


# ========================================
# TestNormalizeColumnName (7 tests)
# ========================================


class TestNormalizeColumnName:
    """Column synonym detection and normalization."""

    def test_exact_canonical_key(self):
        assert normalize_column_name("nom") == "nom"
        assert normalize_column_name("code_postal") == "code_postal"
        assert normalize_column_name("delivery_code") == "delivery_code"

    def test_synonym_detection(self):
        assert normalize_column_name("PRM") == "delivery_code"
        assert normalize_column_name("pdl") == "delivery_code"
        assert normalize_column_name("PCE") == "delivery_code"
        assert normalize_column_name("point_livraison") == "delivery_code"
        assert normalize_column_name("zip") == "code_postal"
        assert normalize_column_name("city") == "ville"

    def test_accent_removal(self):
        assert normalize_column_name("énergie") == "energy_type"
        assert normalize_column_name("numéro_série") == "numero_serie"
        assert normalize_column_name("catégorie") == "type"

    def test_whitespace_and_hyphens(self):
        assert normalize_column_name("  code postal  ") == "code_postal"
        assert normalize_column_name("code-postal") == "code_postal"
        assert normalize_column_name("type  compteur") == "type_compteur"

    def test_bom_stripping(self):
        assert normalize_column_name("\ufeffnom") == "nom"

    def test_case_insensitive(self):
        assert normalize_column_name("NOM") == "nom"
        assert normalize_column_name("Surface_M2") == "surface_m2"
        assert normalize_column_name("SIRET") == "siret"

    def test_unknown_column_passthrough(self):
        result = normalize_column_name("my_custom_column")
        assert result == "my_custom_column"


# ========================================
# TestMapHeaders (4 tests)
# ========================================


class TestMapHeaders:
    """Header mapping from raw CSV/Excel headers to canonical keys."""

    def test_perfect_headers(self):
        headers = ["nom", "adresse", "code_postal", "ville"]
        mapping, warnings = map_headers(headers)
        assert mapping == {"nom": "nom", "adresse": "adresse", "code_postal": "code_postal", "ville": "ville"}
        assert len(warnings) == 0

    def test_synonym_headers(self):
        headers = ["name", "address", "zip", "city", "PRM"]
        mapping, warnings = map_headers(headers)
        assert mapping["name"] == "nom"
        assert mapping["PRM"] == "delivery_code"
        assert mapping["zip"] == "code_postal"
        assert len(warnings) == 0

    def test_unknown_headers_generate_warnings(self):
        headers = ["nom", "custom_field", "another_unknown"]
        mapping, warnings = map_headers(headers)
        assert len(mapping) == 1  # only "nom"
        assert len(warnings) == 2
        assert any("custom_field" in w["header"] for w in warnings)

    def test_duplicate_mapping_warns(self):
        headers = ["nom", "name"]  # both map to "nom"
        mapping, warnings = map_headers(headers)
        assert len(mapping) == 1  # first wins
        assert len(warnings) == 1  # second warned as duplicate


# ========================================
# TestDetectDelimiter (3 tests)
# ========================================


class TestDetectDelimiter:
    def test_semicolon(self):
        assert detect_delimiter("nom;adresse;ville") == ";"

    def test_comma(self):
        assert detect_delimiter("nom,adresse,ville") == ","

    def test_tab(self):
        assert detect_delimiter("nom\tadresse\tville") == "\t"


# ========================================
# TestDetectEncoding (3 tests)
# ========================================


class TestDetectEncoding:
    def test_utf8(self):
        assert detect_encoding("Hello".encode("utf-8")) == "utf-8"

    def test_utf8_bom(self):
        assert detect_encoding(b"\xef\xbb\xbfHello") == "utf-8-sig"

    def test_latin1_fallback(self):
        # Latin-1 bytes that are invalid UTF-8
        data = "Résumé".encode("latin-1")
        assert detect_encoding(data) == "latin-1"


# ========================================
# TestNormalizeRows (4 tests)
# ========================================


class TestNormalizeRows:
    """Value normalization in parsed rows."""

    def test_postal_code_padding(self):
        mapping = {"cp": "code_postal"}
        rows = [{"cp": "1234"}]
        result, _ = normalize_rows(rows, mapping)
        assert result[0]["code_postal"] == "01234"

    def test_delivery_code_to_meter_id(self):
        mapping = {"PRM": "delivery_code"}
        rows = [{"PRM": "12345678901234"}]
        result, _ = normalize_rows(rows, mapping)
        assert "meter_id" in result[0]
        assert result[0]["meter_id"] == "12345678901234"
        assert "delivery_code" not in result[0]

    def test_energy_type_normalization(self):
        mapping = {"energie": "energy_type"}
        rows = [{"energie": "Electricite"}, {"energie": "GAZ naturel"}]
        result, _ = normalize_rows(rows, mapping)
        assert result[0]["energy_type"] == "elec"
        assert result[1]["energy_type"] == "gaz"

    def test_siren_from_siret(self):
        mapping = {"siret": "siret"}
        rows = [{"siret": "44306184100015"}]
        result, _ = normalize_rows(rows, mapping)
        assert result[0]["siret"] == "44306184100015"
        assert result[0]["siren"] == "443061841"


# ========================================
# TestGenerateTemplate (4 tests)
# ========================================


class TestGenerateTemplate:
    """CSV and Excel template generation."""

    def test_csv_template_has_all_columns(self):
        content = generate_csv_template()
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), delimiter=";")
        headers = next(reader)
        expected_keys = [col["key"] for col in CANONICAL_COLUMNS]
        assert headers == expected_keys

    def test_csv_template_has_example_row(self):
        content = generate_csv_template()
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), delimiter=";")
        next(reader)  # headers
        example = next(reader)
        assert len(example) == len(CANONICAL_COLUMNS)
        assert example[0] == "Mairie Principale"  # nom example

    @pytest.mark.skipif(not _has_openpyxl, reason="openpyxl not installed")
    def test_xlsx_template_has_patrimoine_sheet(self):
        content = generate_xlsx_template()
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        assert "Patrimoine" in wb.sheetnames

    @pytest.mark.skipif(not _has_openpyxl, reason="openpyxl not installed")
    def test_xlsx_template_has_aide_sheet(self):
        content = generate_xlsx_template()
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        assert "Aide" in wb.sheetnames
        ws = wb["Aide"]
        # Check synonyms are listed
        assert ws.cell(row=1, column=4).value == "Synonymes acceptes"


# ========================================
# TestTemplateEndpoints (4 tests)
# ========================================


class TestTemplateEndpoints:
    """API endpoints: template download + columns metadata."""

    def test_template_csv_download(self, client):
        response = client.get("/api/patrimoine/import/template?format=csv")
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "template_patrimoine.csv" in response.headers["content-disposition"]
        # Check content has headers
        text = response.content.decode("utf-8-sig")
        assert "nom" in text
        assert "delivery_code" in text

    @pytest.mark.skipif(not _has_openpyxl, reason="openpyxl not installed")
    def test_template_xlsx_download(self, client):
        response = client.get("/api/patrimoine/import/template?format=xlsx")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    @pytest.mark.skipif(not _has_openpyxl, reason="openpyxl not installed")
    def test_template_default_xlsx(self, client):
        response = client.get("/api/patrimoine/import/template")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]

    def test_template_columns_metadata(self, client):
        response = client.get("/api/patrimoine/import/template/columns")
        assert response.status_code == 200
        data = response.json()
        assert "columns" in data
        assert len(data["columns"]) == 14
        # Check required field
        nom_col = next(c for c in data["columns"] if c["key"] == "nom")
        assert nom_col["required"] is True


# ========================================
# TestStagingRowsEndpoint (4 tests)
# ========================================


class TestStagingRowsEndpoint:
    """GET /staging/{id}/rows — paginated, searchable."""

    def test_rows_returns_sites_and_compteurs(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, sites, compteurs = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/rows")
        assert response.status_code == 200
        data = response.json()
        assert data["total"] == 2
        assert len(data["rows"]) == 2
        # First site should have 2 compteurs
        site_alpha = next(r for r in data["rows"] if r["nom"] == "Site Alpha")
        assert len(site_alpha["compteurs"]) == 2

    def test_rows_search_filter(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, _, _ = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/rows?q=Alpha")
        data = response.json()
        assert data["total"] == 1
        assert data["rows"][0]["nom"] == "Site Alpha"

    def test_rows_pagination(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, _, _ = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/rows?page=1&page_size=1")
        data = response.json()
        assert data["total"] == 2
        assert len(data["rows"]) == 1  # page_size=1

    def test_rows_404_unknown_batch(self, client, db_session):
        # Need an org so scope resolution succeeds (demo fallback)
        org = Organisation(nom="Tmp Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()
        response = client.get("/api/patrimoine/staging/9999/rows")
        assert response.status_code == 404


# ========================================
# TestStagingIssuesEndpoint (3 tests)
# ========================================


class TestStagingIssuesEndpoint:
    """GET /staging/{id}/issues — severity filter."""

    def test_issues_returns_findings(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, sites, _ = _create_batch_with_data(db_session, org.id)
        # Add a quality finding
        f = QualityFinding(
            batch_id=batch.id,
            staging_site_id=sites[0].id,
            rule_id="dup_site_address",
            severity=QualityRuleSeverity.BLOCKING,
            evidence_json='{"detail": "test"}',
        )
        db_session.add(f)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/issues")
        data = response.json()
        assert data["total"] == 1
        assert data["issues"][0]["rule_id"] == "dup_site_address"

    def test_issues_severity_filter(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, sites, _ = _create_batch_with_data(db_session, org.id)
        f1 = QualityFinding(
            batch_id=batch.id, staging_site_id=sites[0].id, rule_id="dup", severity=QualityRuleSeverity.BLOCKING
        )
        f2 = QualityFinding(
            batch_id=batch.id, staging_site_id=sites[1].id, rule_id="incomplete", severity=QualityRuleSeverity.WARNING
        )
        db_session.add_all([f1, f2])
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/issues?severity=warning")
        data = response.json()
        assert data["total"] == 1
        assert data["issues"][0]["severity"] == "warning"

    def test_issues_404_unknown_batch(self, client, db_session):
        # Need an org so scope resolution succeeds (demo fallback)
        org = Organisation(nom="Tmp Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()
        response = client.get("/api/patrimoine/staging/9999/issues")
        assert response.status_code == 404


# ========================================
# TestBulkFixEndpoint (2 tests)
# ========================================


class TestBulkFixEndpoint:
    """PUT /staging/{id}/fix/bulk — batch corrections."""

    def test_bulk_fix_skip_multiple(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, sites, _ = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.put(
            f"/api/patrimoine/staging/{batch.id}/fix/bulk",
            json={
                "fixes": [
                    {"fix_type": "skip", "params": {"staging_site_id": sites[0].id}},
                    {"fix_type": "skip", "params": {"staging_site_id": sites[1].id}},
                ]
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["applied"] == 2
        assert data["total"] == 2

    def test_bulk_fix_empty_list(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, _, _ = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.put(
            f"/api/patrimoine/staging/{batch.id}/fix/bulk",
            json={"fixes": []},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["applied"] == 0


# ========================================
# TestAutofix (4 tests)
# ========================================


class TestAutofix:
    """POST /staging/{id}/autofix — safe auto-corrections."""

    def test_autofix_trims_whitespace(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch = create_staging_batch(
            db_session,
            org_id=org.id,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        s = StagingSite(batch_id=batch.id, row_number=2, nom="  Site Spaces  ", ville="  Paris  ")
        db_session.add(s)
        db_session.commit()

        response = client.post(f"/api/patrimoine/staging/{batch.id}/autofix")
        assert response.status_code == 200
        data = response.json()
        assert data["fixes_applied"] >= 1

        # Verify trimmed
        db_session.refresh(s)
        assert s.nom == "Site Spaces"
        assert s.ville == "Paris"

    def test_autofix_pads_postal_code(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch = create_staging_batch(
            db_session,
            org_id=org.id,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        s = StagingSite(batch_id=batch.id, row_number=2, nom="Site CP", code_postal="1234")
        db_session.add(s)
        db_session.commit()

        response = client.post(f"/api/patrimoine/staging/{batch.id}/autofix")
        assert response.status_code == 200

        db_session.refresh(s)
        assert s.code_postal == "01234"

    def test_autofix_normalizes_compteur_type(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch = create_staging_batch(
            db_session,
            org_id=org.id,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        s = StagingSite(batch_id=batch.id, row_number=2, nom="Site Norm")
        db_session.add(s)
        db_session.flush()

        c = StagingCompteur(batch_id=batch.id, staging_site_id=s.id, numero_serie="CPT-NORM", type_compteur="Elec")
        db_session.add(c)
        db_session.commit()

        response = client.post(f"/api/patrimoine/staging/{batch.id}/autofix")
        assert response.status_code == 200

        db_session.refresh(c)
        assert c.type_compteur == "electricite"

    def test_autofix_skips_empty_compteurs(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch = create_staging_batch(
            db_session,
            org_id=org.id,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        s = StagingSite(batch_id=batch.id, row_number=2, nom="Site Empty")
        db_session.add(s)
        db_session.flush()

        c = StagingCompteur(batch_id=batch.id, staging_site_id=s.id)
        db_session.add(c)
        db_session.commit()

        response = client.post(f"/api/patrimoine/staging/{batch.id}/autofix")
        assert response.status_code == 200

        db_session.refresh(c)
        assert c.skip is True


# ========================================
# TestStagingResultEndpoint (2 tests)
# ========================================


class TestStagingResultEndpoint:
    """GET /staging/{id}/result — post-activation summary."""

    def test_result_after_activation(self, client, db_session):
        org, ej, pf = _create_org(db_session)
        batch, _, _ = _create_batch_with_data(db_session, org.id)
        run_quality_gate(db_session, batch.id)
        activate_batch(db_session, batch.id, pf.id)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/result")
        assert response.status_code == 200
        data = response.json()
        assert data["batch_id"] == batch.id
        assert data["status"] == "applied"
        assert "activation" in data
        assert data["activation"]["sites_created"] == 2

    def test_result_404_unknown_batch(self, client, db_session):
        # Need an org so scope resolution succeeds (demo fallback)
        org = Organisation(nom="Tmp Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()
        response = client.get("/api/patrimoine/staging/9999/result")
        assert response.status_code == 404


# ========================================
# TestExportReportEndpoint (2 tests)
# ========================================


class TestExportReportEndpoint:
    """GET /staging/{id}/export/report.csv — CSV export."""

    def test_export_report_csv(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        batch, sites, _ = _create_batch_with_data(db_session, org.id)
        db_session.commit()

        response = client.get(f"/api/patrimoine/staging/{batch.id}/export/report.csv")
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]

        text = response.content.decode("utf-8-sig")
        lines = text.strip().split("\n")
        assert len(lines) == 3  # header + 2 sites
        assert "Site Alpha" in lines[1]

    def test_export_report_404_unknown(self, client, db_session):
        # Need an org so scope resolution succeeds (demo fallback)
        org = Organisation(nom="Tmp Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()
        response = client.get("/api/patrimoine/staging/9999/export/report.csv")
        assert response.status_code == 404


# ========================================
# TestCSVImportWithSynonyms (3 tests)
# ========================================


class TestCSVImportWithSynonyms:
    """End-to-end: CSV with synonym columns → staging correctly populated."""

    def test_delivery_code_column_maps_to_meter_id(self, db_session):
        batch = create_staging_batch(
            db_session,
            org_id=None,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        csv_content = (
            "nom,adresse,code_postal,ville,delivery_code,type_compteur\n"
            "Bureau Paris,10 rue Paix,75002,Paris,12345678901234,electricite\n"
        ).encode("utf-8")

        result = import_csv_to_staging(db_session, batch.id, csv_content)
        assert result["sites_count"] == 1
        assert result["compteurs_count"] == 1

        cpt = (
            db_session.query(StagingCompteur)
            .filter(
                StagingCompteur.batch_id == batch.id,
            )
            .first()
        )
        assert cpt.meter_id == "12345678901234"

    def test_energy_type_infers_type_compteur(self, db_session):
        batch = create_staging_batch(
            db_session,
            org_id=None,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        csv_content = ("nom,numero_serie,energy_type\nSite A,CPT-001,elec\nSite B,CPT-002,gaz\n").encode("utf-8")

        import_csv_to_staging(db_session, batch.id, csv_content)
        compteurs = (
            db_session.query(StagingCompteur)
            .filter(
                StagingCompteur.batch_id == batch.id,
            )
            .order_by(StagingCompteur.id)
            .all()
        )

        assert compteurs[0].type_compteur == "electricite"
        assert compteurs[1].type_compteur == "gaz"

    def test_semicolon_csv_import(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        csv_content = "nom;adresse;code_postal;ville\nSite SC;1 rue Test;75001;Paris\n"
        response = client.post(
            "/api/patrimoine/staging/import",
            files={"file": ("test.csv", csv_content.encode(), "text/csv")},
            params={"mode": "express"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["sites_count"] == 1


# ========================================
# TestActivationIdempotent (2 tests)
# ========================================


class TestActivationIdempotent:
    """Validate → Activate twice = no duplicates."""

    def test_double_activation_no_duplicates(self, db_session):
        org, ej, pf = _create_org(db_session)
        batch, _, _ = _create_batch_with_data(db_session, org.id)
        run_quality_gate(db_session, batch.id)

        result1 = activate_batch(db_session, batch.id, pf.id)
        assert result1["sites_created"] == 2

        result2 = activate_batch(db_session, batch.id, pf.id)
        assert "already applied" in result2["detail"].lower()

        # Only 2 real sites
        real_sites = db_session.query(Site).filter(Site.portefeuille_id == pf.id).all()
        assert len(real_sites) == 2

    def test_activation_creates_delivery_points(self, db_session):
        org, ej, pf = _create_org(db_session)

        batch = create_staging_batch(
            db_session,
            org_id=org.id,
            user_id=None,
            source_type=ImportSourceType.CSV,
            mode="import",
        )
        ss = StagingSite(
            batch_id=batch.id, row_number=2, nom="DP Site", adresse="1 rue DP", code_postal="75001", ville="Paris"
        )
        db_session.add(ss)
        db_session.flush()

        sc = StagingCompteur(
            batch_id=batch.id,
            staging_site_id=ss.id,
            numero_serie="DP-001",
            meter_id="98765432109876",
            type_compteur="electricite",
        )
        db_session.add(sc)
        db_session.flush()

        run_quality_gate(db_session, batch.id)
        result = activate_batch(db_session, batch.id, pf.id)

        assert result["delivery_points_created"] >= 1

        dp = (
            db_session.query(DeliveryPoint)
            .filter(
                DeliveryPoint.code == "98765432109876",
            )
            .first()
        )
        assert dp is not None
        assert dp.energy_type == DeliveryPointEnergyType.ELEC

        cpt = (
            db_session.query(Compteur)
            .filter(
                Compteur.meter_id == "98765432109876",
            )
            .first()
        )
        assert cpt.delivery_point_id == dp.id


# ========================================
# TestImportMappingEndpoint (2 tests)
# ========================================


class TestImportMappingEndpoint:
    """POST /staging/import returns mapping info for CSV files."""

    def test_import_returns_mapping_info(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        csv_content = "name,address,zip,city\nSite M,1 rue Map,75001,Paris\n"
        response = client.post(
            "/api/patrimoine/staging/import",
            files={"file": ("mapping_test.csv", csv_content.encode(), "text/csv")},
            params={"mode": "express"},
        )
        assert response.status_code == 200
        data = response.json()
        assert "mapping" in data
        mapping = data["mapping"]
        assert "mapping" in mapping
        assert mapping["encoding"] in ("utf-8", "utf-8-sig")

    def test_import_duplicate_detection(self, client, db_session):
        org = Organisation(nom="API Org", type_client="bureau", actif=True)
        db_session.add(org)
        db_session.commit()

        csv_content = "nom,ville\nSite Dup,Paris\n"
        # First import
        r1 = client.post(
            "/api/patrimoine/staging/import",
            files={"file": ("dup.csv", csv_content.encode(), "text/csv")},
            params={"mode": "express"},
        )
        assert r1.status_code == 200
        assert r1.json()["duplicate"] is False

        # Second import with same content
        r2 = client.post(
            "/api/patrimoine/staging/import",
            files={"file": ("dup.csv", csv_content.encode(), "text/csv")},
            params={"mode": "express"},
        )
        assert r2.status_code == 200
        assert r2.json()["duplicate"] is True
