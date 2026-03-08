"""
PROMEOS - Import Mapping Service (WORLD CLASS)
Column synonym detection, normalization, auto-detect delimiter/encoding.
FR/EN column synonym resolution for CSV/Excel import.
Maps user-provided headers to canonical column names used by the staging pipeline.
"""

import io
import csv
import re
from typing import Dict, List, Tuple, Optional

# ========================================
# Canonical columns (order matters for template)
# ========================================

CANONICAL_COLUMNS = [
    {"key": "nom", "label": "Nom du site", "required": True, "example": "Mairie Principale"},
    {"key": "adresse", "label": "Adresse", "required": False, "example": "1 place de la Republique"},
    {"key": "code_postal", "label": "Code postal", "required": False, "example": "75001"},
    {"key": "ville", "label": "Ville", "required": False, "example": "Paris"},
    {"key": "surface_m2", "label": "Surface (m2)", "required": False, "example": "1200"},
    {"key": "type", "label": "Type de site", "required": False, "example": "bureau"},
    {"key": "naf_code", "label": "Code NAF", "required": False, "example": "84.11Z"},
    {"key": "siren", "label": "SIREN", "required": False, "example": "443061841"},
    {"key": "siret", "label": "SIRET", "required": False, "example": "44306184100015"},
    {"key": "energy_type", "label": "Type energie", "required": False, "example": "elec"},
    {"key": "delivery_code", "label": "Code PRM/PDL/PCE", "required": False, "example": "12345678901234"},
    {"key": "numero_serie", "label": "N° serie compteur", "required": False, "example": "CPT-001"},
    {"key": "type_compteur", "label": "Type compteur", "required": False, "example": "electricite"},
    {"key": "puissance_kw", "label": "Puissance (kW)", "required": False, "example": "36"},
    # Hiérarchie multi-entité (optionnelles — Step 20)
    {
        "key": "siren_entite",
        "label": "SIREN Entité",
        "required": False,
        "example": "123456789",
        "description": "SIREN de l'entité juridique (9 chiffres). Si fourni, le site est rattaché à cette entité.",
    },
    {
        "key": "nom_entite",
        "label": "Nom Entité",
        "required": False,
        "example": "Filiale Nord",
        "description": "Nom de l'entité juridique. Utilisé avec siren_entite pour créer/réutiliser l'entité.",
    },
    {
        "key": "portefeuille",
        "label": "Portefeuille",
        "required": False,
        "example": "Retail IDF",
        "description": "Nom du portefeuille. Si fourni, le site est rattaché à ce portefeuille.",
    },
    # Bâtiment (optionnelles — Step 20)
    {
        "key": "batiment_nom",
        "label": "Nom Bâtiment",
        "required": False,
        "example": "Tour A",
        "description": "Nom du bâtiment. Si fourni, un bâtiment est créé et rattaché au site.",
    },
    {
        "key": "batiment_surface_m2",
        "label": "Surface Bâtiment (m²)",
        "required": False,
        "example": "800",
        "description": "Surface du bâtiment en m².",
    },
    {
        "key": "batiment_annee_construction",
        "label": "Année Construction",
        "required": False,
        "example": "1995",
        "description": "Année de construction du bâtiment.",
    },
    {
        "key": "batiment_cvc_power_kw",
        "label": "Puissance CVC (kW)",
        "required": False,
        "example": "120",
        "description": "Puissance CVC du bâtiment en kW (utile pour BACS).",
    },
]

# Set of all canonical column keys (used by patrimoine CRUD + staging pipeline)
CANONICAL_COLUMN_KEYS = {
    "nom",
    "adresse",
    "code_postal",
    "ville",
    "surface_m2",
    "type",
    "naf_code",
    "siret",
    "numero_serie",
    "meter_id",
    "type_compteur",
    "puissance_kw",
    "region",
    "nombre_employes",
    # Contract columns
    "fournisseur",
    "energie",
    "prix_kwh",
    "date_debut",
    "date_fin",
    "preavis_jours",
    "abonnement_mensuel",
    # Multi-entité / bâtiment (Step 20)
    "siren_entite",
    "nom_entite",
    "portefeuille",
    "batiment_nom",
    "batiment_surface_m2",
    "batiment_annee_construction",
    "batiment_cvc_power_kw",
}

# ========================================
# Column synonym dictionary (canonical → list of synonyms)
# Used by normalize_column_name, map_headers, template generation
# ========================================

_SYNONYMS = {
    # Site fields
    "nom": ["nom", "name", "site_name", "nom_site", "site", "designation", "libelle", "label", "intitule"],
    "adresse": ["adresse", "address", "addr", "rue", "street", "adresse_site", "adresse_postale"],
    "code_postal": [
        "code_postal",
        "cp",
        "postal_code",
        "zip",
        "zipcode",
        "zip_code",
        "code_post",
        "codepostal",
        "code",
    ],
    "ville": ["ville", "city", "commune", "localite", "town"],
    "region": ["region", "departement", "dept"],
    "surface_m2": ["surface_m2", "surface", "area", "superficie", "m2", "surface_totale", "area_m2", "sup_m2"],
    "type": ["type", "type_site", "usage", "categorie", "category", "activite", "site_type", "nature", "typologie"],
    "naf_code": ["naf_code", "naf", "code_naf", "ape", "code_ape"],
    "siren": ["siren", "n_siren", "num_siren"],
    "siret": ["siret", "n_siret", "num_siret", "siret_site", "no_siret", "numero_siret"],
    "nombre_employes": ["nombre_employes", "employes", "effectif", "nb_employes", "employees", "headcount"],
    # Meter / delivery point fields
    "energy_type": ["energy_type", "energie", "energy", "type_energie", "fluide", "fluid", "vecteur"],
    "delivery_code": [
        "delivery_code",
        "meter_id",
        "prm",
        "pdl",
        "pce",
        "point_livraison",
        "code_prm",
        "code_pdl",
        "code_pce",
        "num_prm",
        "num_pdl",
        "num_pce",
        "prm_pdl",
        "prm_pce",
        "numero_pdl",
        "numero_prm",
        "point_de_livraison",
        "pdl_pce",
        "identifiant_compteur",
    ],
    "numero_serie": [
        "numero_serie",
        "serial",
        "serial_number",
        "n_serie",
        "num_serie",
        "compteur",
        "num_compteur",
        "numero_compteur",
        "no_serie",
    ],
    "type_compteur": ["type_compteur", "meter_type", "type_meter", "compteur_type"],
    "puissance_kw": [
        "puissance_kw",
        "puissance",
        "power",
        "power_kw",
        "kw",
        "kva",
        "puissance_souscrite",
        "subscribed_power",
        "puissance_souscrite_kw",
    ],
    # Contract-related synonyms
    "fournisseur": ["fournisseur", "supplier", "supplier_name", "nom_fournisseur", "prestataire"],
    "prix_kwh": ["prix_kwh", "prix_eur_kwh", "prix_unitaire", "tarif", "tarif_kwh", "price", "unit_price"],
    "date_debut": ["date_debut", "debut_contrat", "start_date", "debut"],
    "date_fin": ["date_fin", "fin_contrat", "end_date", "echeance", "fin"],
    "preavis_jours": ["preavis_jours", "preavis", "notice_days", "notice_period"],
    "abonnement_mensuel": ["abonnement_mensuel", "abonnement", "fixed_fee", "abo_mensuel"],
    # Multi-entité / bâtiment (Step 20)
    "siren_entite": ["siren_entite", "siren_filiale", "siren_entity", "entity_siren"],
    "nom_entite": ["nom_entite", "entite", "entity_name", "filiale", "nom_filiale"],
    "portefeuille": ["portefeuille", "portfolio", "groupe_sites", "group"],
    "batiment_nom": ["batiment_nom", "building_name", "batiment", "building", "nom_batiment"],
    "batiment_surface_m2": ["batiment_surface_m2", "building_surface", "surface_batiment"],
    "batiment_annee_construction": [
        "batiment_annee_construction",
        "annee_construction",
        "year_built",
        "construction_year",
    ],
    "batiment_cvc_power_kw": ["batiment_cvc_power_kw", "cvc_power", "hvac_power", "puissance_cvc"],
}

# Build reverse lookup: lowered synonym → canonical key
_REVERSE_MAP = {}
for canonical, synonyms in _SYNONYMS.items():
    for syn in synonyms:
        _REVERSE_MAP[syn.lower().strip()] = canonical

# Flat synonym dict (synonym → canonical) used by normalize_header
_FLAT_SYNONYMS: Dict[str, str] = {}
for _canonical, _syn_list in _SYNONYMS.items():
    for _syn in _syn_list:
        _FLAT_SYNONYMS[_syn.lower().strip()] = _canonical


# ========================================
# Type value normalization (WORLD CLASS)
# ========================================

_TYPE_SITE_SYNONYMS: Dict[str, str] = {
    # bureau
    "bureau": "bureau",
    "bureaux": "bureau",
    "office": "bureau",
    "tertiaire": "bureau",
    # commerce
    "commerce": "commerce",
    "magasin": "magasin",
    "boutique": "commerce",
    "retail": "commerce",
    "supermarche": "commerce",
    "hypermarche": "commerce",
    # entrepot
    "entrepot": "entrepot",
    "logistique": "entrepot",
    "warehouse": "entrepot",
    "stockage": "entrepot",
    "depot": "entrepot",
    # usine
    "usine": "usine",
    "industrie": "usine",
    "production": "usine",
    "factory": "usine",
    "industriel": "usine",
    # hotel
    "hotel": "hotel",
    "hotellerie": "hotel",
    "hebergement": "hotel",
    # sante
    "sante": "sante",
    "hopital": "sante",
    "clinique": "sante",
    "ehpad": "sante",
    "medico_social": "sante",
    # enseignement
    "enseignement": "enseignement",
    "ecole": "enseignement",
    "lycee": "enseignement",
    "college": "enseignement",
    "universite": "enseignement",
    "education": "enseignement",
    # copropriete
    "copropriete": "copropriete",
    "copro": "copropriete",
    "syndic": "copropriete",
    "residence": "copropriete",
    "immeuble": "copropriete",
    # collectivite
    "collectivite": "collectivite",
    "mairie": "collectivite",
    "administration": "collectivite",
    "public": "collectivite",
    # logement social
    "logement_social": "logement_social",
    "hlm": "logement_social",
    "social": "logement_social",
    "bailleur": "logement_social",
}

_TYPE_COMPTEUR_SYNONYMS: Dict[str, str] = {
    "electricite": "electricite",
    "elec": "electricite",
    "electricity": "electricite",
    "electrique": "electricite",
    "courant": "electricite",
    "gaz": "gaz",
    "gas": "gaz",
    "naturel": "gaz",
    "eau": "eau",
    "water": "eau",
}


# ========================================
# Column name normalization (v9 advanced)
# ========================================


def normalize_column_name(raw: str) -> str:
    """Normalize a raw column header to canonical form.

    1. Strip whitespace, BOM
    2. Lowercase
    3. Replace spaces/hyphens with underscores
    4. Remove accents (simple)
    5. Lookup in synonym dict
    """
    cleaned = raw.strip().strip("\ufeff").lower()
    cleaned = re.sub(r"[\s\-]+", "_", cleaned)
    # Simple accent removal
    for a, b in [
        ("é", "e"),
        ("è", "e"),
        ("ê", "e"),
        ("ë", "e"),
        ("à", "a"),
        ("â", "a"),
        ("ù", "u"),
        ("û", "u"),
        ("ô", "o"),
        ("î", "i"),
        ("ï", "i"),
        ("ç", "c"),
    ]:
        cleaned = cleaned.replace(a, b)
    # Remove trailing/leading underscores
    cleaned = cleaned.strip("_")

    return _REVERSE_MAP.get(cleaned, cleaned)


# ========================================
# Encoding & delimiter detection (v9)
# ========================================


def detect_delimiter(first_line: str) -> str:
    """Auto-detect CSV delimiter from first line."""
    if "\t" in first_line:
        return "\t"
    semicolons = first_line.count(";")
    commas = first_line.count(",")
    return ";" if semicolons > commas else ","


def detect_encoding(raw_bytes: bytes) -> str:
    """Detect encoding from raw bytes. Returns codec name."""
    if raw_bytes[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    # Try UTF-8 first
    try:
        raw_bytes[:4096].decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    # Fallback: Latin-1 (always succeeds)
    return "latin-1"


# ========================================
# Header mapping (v9 template-aware)
# ========================================


def map_headers(raw_headers: List[str]) -> Tuple[dict, List[dict]]:
    """Map raw CSV/Excel headers to canonical columns.

    Returns:
        mapping: {raw_header: canonical_key} for recognized columns
        warnings: list of {header, message} for unrecognized columns
    """
    mapping = {}
    warnings = []
    used_canonical = set()

    for raw in raw_headers:
        canonical = normalize_column_name(raw)
        if canonical in {col["key"] for col in CANONICAL_COLUMNS} or canonical in CANONICAL_COLUMN_KEYS:
            if canonical in used_canonical:
                warnings.append(
                    {
                        "header": raw,
                        "message": f"Duplicate mapping to '{canonical}' — column ignored",
                    }
                )
            else:
                mapping[raw] = canonical
                used_canonical.add(canonical)
        else:
            warnings.append(
                {
                    "header": raw,
                    "message": f"Column '{raw}' not recognized — will be ignored",
                }
            )

    return mapping, warnings


def normalize_rows(
    raw_rows: List[dict],
    header_mapping: dict,
) -> Tuple[List[dict], List[dict]]:
    """Apply header mapping + value normalization to parsed rows.

    Returns (normalized_rows, row_warnings).
    """
    normalized = []
    row_warnings = []

    for i, raw_row in enumerate(raw_rows, start=2):  # row 1 = header
        row = {}
        for raw_key, value in raw_row.items():
            canonical = header_mapping.get(raw_key)
            if not canonical:
                continue
            row[canonical] = _normalize_value(canonical, value)

        # delivery_code → meter_id (backward compat with existing pipeline)
        if "delivery_code" in row and "meter_id" not in row:
            row["meter_id"] = row.pop("delivery_code")
        elif "delivery_code" in row:
            row.pop("delivery_code")

        # siren extraction from siret if siren missing
        if not row.get("siren") and row.get("siret") and len(str(row["siret"])) >= 9:
            row["siren"] = str(row["siret"])[:9]

        normalized.append(row)

    return normalized, row_warnings


def _normalize_value(key: str, value) -> Optional[str]:
    """Normalize a single cell value."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "null", "n/a", "na", "nd", "-"):
        return None

    # Type-specific normalization
    if key == "code_postal":
        s = s.replace(" ", "").zfill(5)[:5]
    elif key in ("surface_m2", "puissance_kw"):
        s = s.replace(",", ".").replace(" ", "")
    elif key in ("siren", "siret", "delivery_code", "meter_id", "siren_entite"):
        s = s.replace(" ", "").replace("-", "").replace(".", "")
    elif key == "energy_type":
        s = _normalize_energy_type(s)
    elif key == "type_compteur":
        s = _normalize_compteur_type(s)

    return s if s else None


def _normalize_energy_type(raw: str) -> str:
    """Normalize energy type string."""
    low = raw.lower()
    if any(k in low for k in ("elec", "elect", "prm", "pdl")):
        return "elec"
    if any(k in low for k in ("gaz", "gas", "pce")):
        return "gaz"
    if any(k in low for k in ("eau", "water")):
        return "eau"
    return raw


def _normalize_compteur_type(raw: str) -> str:
    """Normalize compteur type string."""
    low = raw.lower()
    if any(k in low for k in ("elec", "elect")):
        return "electricite"
    if any(k in low for k in ("gaz", "gas")):
        return "gaz"
    if any(k in low for k in ("eau", "water")):
        return "eau"
    return raw


# ========================================
# Template generation (v9)
# ========================================


def generate_csv_template() -> bytes:
    """Generate official CSV import template with headers + example row."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([col["key"] for col in CANONICAL_COLUMNS])
    writer.writerow([col["example"] for col in CANONICAL_COLUMNS])
    return output.getvalue().encode("utf-8-sig")


def generate_xlsx_template() -> bytes:
    """Generate official Excel import template with headers, example, and Aide sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl required for Excel template generation")

    wb = Workbook()

    # ── Main sheet: Patrimoine ──
    ws = wb.active
    ws.title = "Patrimoine"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    required_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    example_font = Font(italic=True, color="6B7280")
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for col_idx, col_def in enumerate(CANONICAL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["key"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col_def["key"]) + 4, 14)

    # Example row
    for col_idx, col_def in enumerate(CANONICAL_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_def["example"])
        cell.font = example_font
        cell.border = thin_border
        if col_def["required"]:
            cell.fill = required_fill

    # Second example row
    examples_2 = {
        "nom": "Ecole Voltaire",
        "adresse": "12 avenue Victor Hugo",
        "code_postal": "69003",
        "ville": "Lyon",
        "surface_m2": "800",
        "type": "enseignement",
        "naf_code": "85.20Z",
        "siren": "217500016",
        "siret": "21750001600015",
        "energy_type": "elec",
        "delivery_code": "98765432109876",
        "numero_serie": "CPT-002",
        "type_compteur": "electricite",
        "puissance_kw": "24",
    }
    for col_idx, col_def in enumerate(CANONICAL_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=examples_2.get(col_def["key"], ""))
        cell.font = example_font
        cell.border = thin_border

    # ── Aide sheet ──
    ws_help = wb.create_sheet("Aide")
    ws_help.column_dimensions["A"].width = 20
    ws_help.column_dimensions["B"].width = 40
    ws_help.column_dimensions["C"].width = 12
    ws_help.column_dimensions["D"].width = 50

    help_headers = ["Colonne", "Description", "Obligatoire", "Synonymes acceptes"]
    for col_idx, h in enumerate(help_headers, start=1):
        cell = ws_help.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, col_def in enumerate(CANONICAL_COLUMNS, start=2):
        ws_help.cell(row=row_idx, column=1, value=col_def["key"])
        ws_help.cell(row=row_idx, column=2, value=col_def["label"])
        ws_help.cell(row=row_idx, column=3, value="Oui" if col_def["required"] else "Non")
        synonyms = _SYNONYMS.get(col_def["key"], [])
        ws_help.cell(row=row_idx, column=4, value=", ".join(synonyms))

    # Notes
    notes_row = len(CANONICAL_COLUMNS) + 3
    ws_help.cell(row=notes_row, column=1, value="NOTES").font = Font(bold=True)
    notes = [
        "- Seule la colonne 'nom' est obligatoire. Les autres sont recommandees.",
        "- Le delimiter CSV est auto-detecte (virgule ou point-virgule).",
        "- Les noms de colonnes sont flexibles: 'CP', 'postal_code', 'zip' → code_postal.",
        "- delivery_code = PRM (elec) ou PCE (gaz), 14 chiffres.",
        "- Les lignes avec le meme 'nom' sont regroupees dans un seul site.",
        "- Formats supportes: CSV (UTF-8), Excel (.xlsx).",
        "- SIRET = 14 chiffres (inclut SIREN = 9 premiers chiffres).",
    ]
    for i, note in enumerate(notes):
        ws_help.cell(row=notes_row + 1 + i, column=1, value=note)

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ========================================
# Public API (WORLD CLASS patrimoine)
# ========================================


def normalize_header(raw_header: str) -> Optional[str]:
    """Map a single raw column header to its canonical name.

    Returns None if no match found.
    """
    key = raw_header.strip().lower().replace("-", "_").replace(" ", "_")
    return _FLAT_SYNONYMS.get(key)


def normalize_headers(raw_headers: List[str]) -> Tuple[Dict[str, str], List[str]]:
    """Map a list of raw CSV/Excel headers to canonical columns.

    Returns:
        (mapping, unmapped) where mapping is {raw_header: canonical_name}
        and unmapped is the list of headers that couldn't be resolved.
    """
    mapping = {}
    unmapped = []
    for h in raw_headers:
        canonical = normalize_header(h)
        if canonical:
            mapping[h] = canonical
        else:
            unmapped.append(h)
    return mapping, unmapped


def normalize_type_site(raw_value: str) -> Optional[str]:
    """Normalize a site type value using FR synonyms."""
    key = raw_value.strip().lower().replace("-", "_").replace(" ", "_")
    return _TYPE_SITE_SYNONYMS.get(key)


def normalize_type_compteur(raw_value: str) -> Optional[str]:
    """Normalize a compteur type value using FR synonyms."""
    key = raw_value.strip().lower().replace("-", "_").replace(" ", "_")
    return _TYPE_COMPTEUR_SYNONYMS.get(key)


def get_mapping_report(raw_headers: List[str]) -> dict:
    """Generate a diagnostic report for header mapping.

    Useful for the PatrimoineWizard to show the user what was recognized.
    """
    mapping, unmapped = normalize_headers(raw_headers)
    canonical_found = set(mapping.values())
    required = {"nom"}
    missing_required = required - canonical_found
    recommended = {"adresse", "code_postal", "ville", "surface_m2", "type"}
    missing_recommended = recommended - canonical_found

    return {
        "mapped": mapping,
        "unmapped": unmapped,
        "canonical_found": sorted(canonical_found),
        "missing_required": sorted(missing_required),
        "missing_recommended": sorted(missing_recommended),
        "coverage_pct": round(len(mapping) / max(len(raw_headers), 1) * 100, 1),
        "is_valid": len(missing_required) == 0,
    }
